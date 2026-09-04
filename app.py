
import os
import io
import re
import json
import time
import hashlib
import hmac
import secrets
import urllib.request
import urllib.error
from datetime import datetime, date, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import streamlit as st
from sqlalchemy import (
    create_engine, MetaData, Table, Column, Integer, String, Text, select, insert, update, text as sql_text
)
from sqlalchemy.exc import IntegrityError

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    from google import genai
except Exception:
    genai = None

APP_VERSION = "V4.7.3 Reliable Navigation Edition"
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = Path(os.getenv("VOC_DB_PATH", str(DATA_DIR / "voc_intelligence.db")))

def _secret(name, default=""):
    try:
        if name in st.secrets:
            return str(st.secrets[name])
    except Exception:
        pass
    return os.getenv(name, default)

DATABASE_URL = _secret("DATABASE_URL", "").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = "postgresql+psycopg://" + DATABASE_URL[len("postgres://"):]
elif DATABASE_URL.startswith("postgresql://"):
    DATABASE_URL = "postgresql+psycopg://" + DATABASE_URL[len("postgresql://"):]

if DATABASE_URL:
    DB_BACKEND = "PostgreSQL · Shared Cloud DB"
    ENGINE = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,
        pool_recycle=300,
        future=True,
    )
else:
    DB_BACKEND = "SQLite · Local Test DB"
    ENGINE = create_engine(
        f"sqlite:///{DB_PATH}",
        connect_args={"check_same_thread": False},
        future=True,
    )

ICON_PATH = BASE_DIR / "assets" / "voc_gothic_logo.png"
PAGE_ICON = str(ICON_PATH) if ICON_PATH.exists() else "📋"

st.set_page_config(
    page_title="VOC Intelligence",
    page_icon=PAGE_ICON,
    layout="wide",
    initial_sidebar_state="expanded",
)

STATUS_OPTIONS = [
    "Open", "In Progress", "Pending", "Over Due",
    "Waiting Customer", "Waiting Internal", "Validation",
    "Hold", "Drop", "Closed"
]
ROLE_OPTIONS = ["Admin", "Editor", "Viewer"]
PRIORITIES = ["Low", "Medium", "High", "Critical"]
PROCESSES = [
    "", "Customer Incoming / Material", "Coating", "Drying/Oven", "Lamination",
    "Slitting", "Die-cut", "Waste Stripping", "Peeling/Release", "Assembly",
    "Reliability Test", "Customer Process", "Other"
]
FAILURES = [
    "", "Appearance", "拉胶 / Stringing", "溢胶 / Adhesive Overflow",
    "残胶 / Residue", "起翘 / Lifting", "Curl", "Bubble", "Wrinkle",
    "Scratch", "Foreign Material", "Release Force Low", "Release Force High",
    "Peel Low", "High-temp Peel", "Dimension", "Other"
]

# Navigation language changes; business form labels remain KO/EN/ZH together.
LANG_OPTIONS = {"한국어": "ko", "中文": "zh"}
NAV = {
    "ko": {
        "subtitle": "한국 본사 ↔ 중국법인 공동 VOC 관리",
        "login_id": "아이디", "login_pw": "비밀번호", "login": "로그인",
        "dashboard": "대시보드", "analysis": "AI VOC 심층 분석",
        "search": "VOC 검색 · 수정", "translate": "한 · 중 · 영 변환",
        "settings": "설정", "logout": "로그아웃", "menu": "메뉴",
        "language": "언어 / Language",
    },
    "zh": {
        "subtitle": "韩国总部 ↔ 中国法人 共同VOC管理",
        "login_id": "用户名", "login_pw": "密码", "login": "登录",
        "dashboard": "仪表盘", "analysis": "AI VOC深度分析",
        "search": "VOC查询 · 修改", "translate": "韩 · 中 · 英转换",
        "settings": "设置", "logout": "退出登录", "menu": "菜单",
        "language": "语言 / Language",
    },
}

UI = {
    "date": "접수일 / Received Date / 接收日期",
    "customer": "고객사 / Customer / 客户",
    "dri": "담당자 / DRI / 负责人",
    "product": "제품명 / Product / 产品",
    "material": "제품 유형 / Material · Product Type / 材料·产品类型",
    "structure": "제품 구조 / Product Structure / 产品结构",
    "lot": "LOT No. / Lot No. / 批次号",
    "quantity": "수량 / Quantity / 数量",
    "process": "공정 / Process / 工序",
    "failure": "불량 분류 / Failure Category / 不良分类",
    "detail": "불량 상세 / Failure Detail / 不良详情",
    "original": "VOC 원문 / Original VOC / VOC原文",
    "spec": "규격 / SPEC / 规格",
    "actual": "실측 / Actual / 实测",
    "judgement": "판정 / Judgement / 判定",
    "priority": "우선순위 / Priority / 优先级",
    "summary": "이슈 요약 / Issue Summary / 问题摘要",
    "condition": "발생 조건 / Occurrence Condition / 发生条件",
    "rate": "불량률 / Defect Rate / 不良率",
    "impact": "고객 영향 / Customer Impact / 客户影响",
    "request": "고객 요청 / Customer Request / 客户要求",
    "response_due": "회신 기한 / Response Due / 回复期限",
    "faca_due": "FACA 기한 / FACA Due / FACA期限",
    "internal": "즉시 내부 조치 / Internal Immediate Action / 内部即时措施",
    "check": "확인 필요 항목 / Required Check Points / 必要确认项目",
    "cause": "AI 원인 가설 / AI Suggested Cause / AI原因假设",
    "missing": "추가 확인 정보 / Missing Information / 待确认信息",
    "ko": "한국어 요약 / Korean Summary / 韩文摘要",
    "zh": "중국어 요약 / Chinese Summary / 中文摘要",
    "en": "영어 요약 / English Summary / 英文摘要",
    "root": "확정 원인 / Confirmed Root Cause / 确认根因",
    "escape": "유출 원인 / Escape Cause / 流出原因",
    "corrective": "개선 조치 / Corrective Action / 改善措施",
    "verification": "검증 결과 / Verification Result / 验证结果",
    "status": "상태 / Status / 状态",
    "faca_no": "FACA No.",
    "picture": "사진·링크 / Picture · Link / 图片·链接",
    "remark": "비고 / Remark / 备注",
}

PROCESS_DISPLAY = {
    "": "",
    "Customer Incoming / Material": "고객 입고/원자재 · Customer Incoming · 客户来料",
    "Coating": "코팅 · Coating · 涂布",
    "Drying/Oven": "건조/오븐 · Drying/Oven · 烘干",
    "Lamination": "합지 · Lamination · 贴合",
    "Slitting": "슬리팅 · Slitting · 分条",
    "Die-cut": "타발 · Die-cut · 模切",
    "Waste Stripping": "배폐 · Waste Stripping · 排废",
    "Peeling/Release": "이형/박리 · Peeling/Release · 离型/剥离",
    "Assembly": "조립 · Assembly · 组装",
    "Reliability Test": "신뢰성 시험 · Reliability Test · 可靠性测试",
    "Customer Process": "고객 공정 · Customer Process · 客户工序",
    "Other": "기타 · Other · 其他",
}
FAILURE_DISPLAY = {
    "": "",
    "Appearance": "외관 · Appearance · 外观",
    "拉胶 / Stringing": "라교 · Stringing · 拉胶",
    "溢胶 / Adhesive Overflow": "접착제 넘침 · Overflow · 溢胶",
    "残胶 / Residue": "잔사 · Residue · 残胶",
    "起翘 / Lifting": "들뜸 · Lifting · 起翘",
    "Curl": "컬 · Curl · 翘曲",
    "Bubble": "기포 · Bubble · 气泡",
    "Wrinkle": "주름 · Wrinkle · 褶皱",
    "Scratch": "스크래치 · Scratch · 划伤",
    "Foreign Material": "이물 · Foreign Material · 异物",
    "Release Force Low": "이형력 낮음 · Low Release Force · 离型力低",
    "Release Force High": "이형력 높음 · High Release Force · 离型力高",
    "Peel Low": "박리력 낮음 · Low Peel · 剥离力低",
    "High-temp Peel": "고온 박리 · High-temp Peel · 高温剥离",
    "Dimension": "치수 · Dimension · 尺寸",
    "Other": "기타 · Other · 其他",
}

VOC_FIELDS = [
    "voc_id", "received_date", "customer", "dri", "product", "material_type",
    "product_structure", "lot_no", "quantity", "process", "failure_category",
    "failure_detail", "voc_original_text", "issue_summary", "summary_ko",
    "summary_zh", "summary_en", "spec", "actual", "judgement",
    "occurrence_condition", "defect_rate", "customer_impact", "priority",
    "customer_request", "response_due", "faca_due", "internal_action_items",
    "ai_suggested_cause", "required_check_points", "missing_information",
    "confirmed_root_cause", "escape_cause", "corrective_action",
    "verification_result", "faca_no", "status", "close_date", "picture_link",
    "remark", "created_by", "created_at", "updated_by", "updated_at",
    "last_update_summary"
]

def tr(key):
    return NAV.get(st.session_state.get("lang", "ko"), NAV["ko"]).get(key, key)

# ---------- Database ----------
metadata = MetaData()

users_table = Table(
    "users", metadata,
    Column("username", String(120), primary_key=True),
    Column("display_name", String(200), nullable=False),
    Column("password_hash", String(256), nullable=False),
    Column("salt", String(128), nullable=False),
    Column("role", String(40), nullable=False, default="Editor"),
    Column("active", Integer, nullable=False, default=1),
    Column("created_at", String(40), nullable=False),
)

voc_table = Table(
    "voc_cases", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("voc_id", String(80), unique=True, nullable=False),
    Column("received_date", String(40)),
    Column("customer", Text), Column("dri", Text), Column("product", Text),
    Column("material_type", Text), Column("product_structure", Text),
    Column("lot_no", Text), Column("quantity", Text), Column("process", Text),
    Column("failure_category", Text), Column("failure_detail", Text),
    Column("voc_original_text", Text), Column("issue_summary", Text),
    Column("summary_ko", Text), Column("summary_zh", Text), Column("summary_en", Text),
    Column("spec", Text), Column("actual", Text), Column("judgement", Text),
    Column("occurrence_condition", Text), Column("defect_rate", Text),
    Column("customer_impact", Text), Column("priority", Text),
    Column("customer_request", Text), Column("response_due", Text),
    Column("faca_due", Text), Column("internal_action_items", Text),
    Column("ai_suggested_cause", Text), Column("required_check_points", Text),
    Column("missing_information", Text), Column("confirmed_root_cause", Text),
    Column("escape_cause", Text), Column("corrective_action", Text),
    Column("verification_result", Text), Column("faca_no", Text),
    Column("status", String(80), default="Open"),
    Column("close_date", String(40)), Column("picture_link", Text),
    Column("remark", Text), Column("created_by", Text), Column("created_at", String(40)),
    Column("updated_by", Text), Column("updated_at", String(40)),
    Column("last_update_summary", Text),
)

audit_table = Table(
    "audit_log", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("voc_id", String(80), nullable=False),
    Column("action", String(80), nullable=False),
    Column("changed_by", String(120), nullable=False),
    Column("changed_at", String(40), nullable=False),
    Column("summary", Text),
)

settings_table = Table(
    "app_settings", metadata,
    Column("key", String(160), primary_key=True),
    Column("value", Text),
)

def hash_password(password, salt_hex):
    salt = bytes.fromhex(salt_hex)
    return hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt, 200_000
    ).hex()

def create_user(conn, username, display_name, password, role):
    username = username.strip()
    salt = secrets.token_bytes(16).hex()
    payload = {
        "username": username,
        "display_name": display_name.strip() or username,
        "password_hash": hash_password(password, salt),
        "salt": salt,
        "role": role,
        "active": 1,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    existing = conn.execute(
        select(users_table.c.username).where(users_table.c.username == username)
    ).first()
    if existing:
        conn.execute(
            update(users_table).where(users_table.c.username == username).values(**payload)
        )
    else:
        conn.execute(insert(users_table).values(**payload))

def init_db():
    metadata.create_all(ENGINE)
    with ENGINE.begin() as conn:
        count = conn.execute(select(users_table.c.username).limit(1)).first()
        if count is None:
            create_user(conn, "admin", "Administrator", "ChangeMe123!", "Admin")

def authenticate(username, password):
    with ENGINE.connect() as conn:
        row = conn.execute(
            select(users_table).where(
                (users_table.c.username == username.strip()) &
                (users_table.c.active == 1)
            )
        ).mappings().first()
    if not row:
        return None
    if hmac.compare_digest(hash_password(password, row["salt"]), row["password_hash"]):
        return dict(row)
    return None

def next_voc_id():
    year = date.today().year
    pattern = f"VOC-{year}-%"
    with ENGINE.connect() as conn:
        rows = conn.execute(
            select(voc_table.c.voc_id).where(voc_table.c.voc_id.like(pattern))
        ).all()
    nums = []
    for r in rows:
        try:
            nums.append(int(r[0].split("-")[-1]))
        except Exception:
            pass
    return f"VOC-{year}-{max(nums, default=0)+1:03d}"

def insert_audit(conn, voc_id, action, user, summary=""):
    conn.execute(insert(audit_table).values(
        voc_id=voc_id,
        action=action,
        changed_by=user,
        changed_at=datetime.now().isoformat(timespec="seconds"),
        summary=summary,
    ))

def save_case(data, username):
    now = datetime.now().isoformat(timespec="seconds")
    payload = {f: data.get(f, "") for f in VOC_FIELDS}
    payload["voc_id"] = payload["voc_id"] or next_voc_id()
    payload["created_by"] = username
    payload["created_at"] = now
    payload["updated_by"] = username
    payload["updated_at"] = now
    payload["last_update_summary"] = "Created"
    with ENGINE.begin() as conn:
        conn.execute(insert(voc_table).values(**payload))
        insert_audit(conn, payload["voc_id"], "CREATE", username, "VOC created")
    return payload["voc_id"]

def update_case(voc_id, updates, username, summary):
    payload = dict(updates)
    payload["updated_by"] = username
    payload["updated_at"] = datetime.now().isoformat(timespec="seconds")
    payload["last_update_summary"] = summary
    if payload.get("status") == "Closed":
        payload["close_date"] = date.today().isoformat()
    with ENGINE.begin() as conn:
        conn.execute(
            update(voc_table).where(voc_table.c.voc_id == voc_id).values(**payload)
        )
        insert_audit(conn, voc_id, "UPDATE", username, summary)

def load_cases():
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            sql_text("SELECT * FROM voc_cases ORDER BY updated_at DESC, id DESC"),
            conn
        )

def load_case(voc_id):
    with ENGINE.connect() as conn:
        row = conn.execute(
            select(voc_table).where(voc_table.c.voc_id == voc_id)
        ).mappings().first()
    return dict(row) if row else None

def load_audit(voc_id):
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            sql_text("""
                SELECT * FROM audit_log
                WHERE voc_id=:voc_id
                ORDER BY id DESC
            """),
            conn,
            params={"voc_id": voc_id},
        )

def change_password(username, new_password):
    salt = secrets.token_bytes(16).hex()
    with ENGINE.begin() as conn:
        conn.execute(
            update(users_table)
            .where(users_table.c.username == username)
            .values(
                password_hash=hash_password(new_password, salt),
                salt=salt,
            )
        )

def list_users_df():
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            sql_text("""
                SELECT username, display_name, role, active, created_at
                FROM users ORDER BY username
            """),
            conn,
        )

def load_all_audit():
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            sql_text("""
                SELECT id, voc_id, action, changed_by, changed_at, summary
                FROM audit_log
                ORDER BY id DESC
            """),
            conn,
        )

def delete_case(voc_id, username):
    """Permanently delete a VOC case and its audit history.

    This action is intentionally restricted at the UI level to Admin users.
    """
    with ENGINE.begin() as conn:
        existing = conn.execute(
            select(voc_table.c.voc_id).where(voc_table.c.voc_id == voc_id)
        ).first()
        if existing is None:
            raise RuntimeError(f"{voc_id} VOC를 찾을 수 없습니다.")

        conn.execute(
            audit_table.delete().where(audit_table.c.voc_id == voc_id)
        )
        conn.execute(
            voc_table.delete().where(voc_table.c.voc_id == voc_id)
        )

def build_excel_backup():
    """Create an in-memory Excel backup of VOC data and audit history."""
    voc_df = load_cases()
    audit_df = load_all_audit()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        voc_df.to_excel(writer, sheet_name="VOC_Master", index=False)
        audit_df.to_excel(writer, sheet_name="Audit_Log", index=False)
        info_df = pd.DataFrame([
            ["Backup Created At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Database Backend", DB_BACKEND],
            ["VOC Records", len(voc_df)],
            ["Audit Records", len(audit_df)],
            ["App Version", APP_VERSION],
        ], columns=["Item", "Value"])
        info_df.to_excel(writer, sheet_name="Backup_Info", index=False)

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = writer.book
        header_fill = PatternFill("solid", fgColor="B89046")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9D9D9")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(bottom=thin)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col_idx, column_cells in enumerate(ws.columns, 1):
                max_len = 0
                for cell in list(column_cells)[:300]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 45))
                ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 38))
            ws.row_dimensions[1].height = 24
    buf.seek(0)
    return buf.getvalue()

# ---------- AI ----------
def clean_json(s):
    s = (s or "").strip()
    s = re.sub(r"^```(?:json)?\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    return s.strip()

def list_to_text(v, numbered=False):
    if isinstance(v, list):
        if numbered:
            return "\n".join(f"{i+1}. {x}" for i, x in enumerate(v) if x)
        return "\n".join(f"- {x}" for x in v if x)
    return v or ""

def heuristic_analysis(text):
    product = ""
    m = re.search(r"\b[A-Z]{2,}[A-Z0-9\-]{3,}\b", text)
    if m:
        product = m.group(0)

    lot = ""
    m = re.search(r"(?:批次号|批次|LOT|Lot|lot)\s*[:：]?\s*([A-Za-z0-9\-]+)", text)
    if m:
        lot = m.group(1)

    qtys = re.findall(r"\d+(?:\.\d+)?\s*MM\s*[*×xX]\s*\d+(?:\.\d+)?\s*M", text, re.I)
    quantity = ", ".join(qtys)

    failure_detail = ""
    if "凹凸点" in text:
        failure_detail = "표면 요철점"
    elif "拉胶" in text:
        failure_detail = "라교 / Stringing"
    elif "起翘" in text:
        failure_detail = "들뜸 / Lifting"

    failure_cat = "Appearance" if ("凹凸点" in text or "外观" in text) else ""
    requests = []
    if "退货" in text:
        requests.append("NG품 반품 처리")
    if "换货" in text or "合格品" in text:
        requests.append("합격품 확인 및 교환")
    if "FACA" in text.upper():
        requests.append("원인 분석 및 FACA 회신")

    return {
        "product": product, "material_type": "", "lot_no": lot, "quantity": quantity,
        "process": "Customer Incoming / Material", "failure_category": failure_cat,
        "failure_detail": failure_detail,
        "issue_summary": re.sub(r"\s+", " ", text.strip())[:300],
        "summary_ko": "", "summary_zh": "", "summary_en": "",
        "spec": "", "actual": "", "judgement": "",
        "occurrence_condition": "", "defect_rate": "",
        "customer_impact": "고객 출하 영향 여부 확인 필요",
        "priority": "High",
        "customer_request": requests,
        "response_due": "", "faca_due": "",
        "internal_action_items": [
            "해당 LOT 재고 Hold 및 선별",
            "동일/인접 LOT 비교",
            "불량 Sample 확보",
            "생산 및 원재료 이력 Trace",
        ],
        "ai_suggested_cause": [
            "원재료 표면 상태",
            "이물/Particle",
            "Coating 균일도",
            "Lamination 또는 권취 압흔",
        ],
        "required_check_points": [
            "OK vs NG Lot 비교",
            "폭/길이 위치별 발생 분포",
            "원재료 LOT 및 공정 이력",
            "불량 Sample 확대 관찰",
        ],
        "missing_information": [
            "전체 불량률",
            "발생 위치 분포",
            "폭별 차이",
            "불량 Sample 및 사진",
        ],
        "analysis_source": "Local / Rule-based"
    }

def build_analysis_prompt(text):
    return f"""
You are a senior NPI / Quality / Technical Sales assistant for functional adhesive tapes,
PSA, conductive fabric tape, die-cutting, lamination, optical/display parts and electronics manufacturing.

Analyze the customer's VOC. The source can mix Simplified Chinese, Korean and English.

Rules:
1. Extract factual fields without inventing missing data.
2. Separate facts from hypotheses.
3. Detect business impact and urgency.
4. Split customer requests into actionable items and deadlines.
5. Compare SPEC vs Actual when mathematically possible.
6. Suggest practical immediate internal actions.
7. Root-cause suggestions MUST remain hypotheses.
8. Generate checkpoints and missing-information questions.
9. ALWAYS generate all three summary fields. summary_ko MUST be Korean, summary_zh MUST be Simplified Chinese, and summary_en MUST be professional English. Never leave these three fields blank when the VOC source contains meaningful content.
10. issue_summary MUST ALWAYS be written in concise professional KOREAN, even when the original VOC is Chinese or English.
11. Other narrative operational fields (failure_detail, occurrence_condition, customer_impact, customer_request, internal_action_items, ai_suggested_cause, required_check_points, missing_information) should also be normalized primarily into Korean.
12. Product codes, LOT, quantities, dates, units and SPEC/Actual values must remain exactly as source.

Return ONLY valid JSON with exactly these keys:
{{
 "product":"",
 "material_type":"",
 "lot_no":"",
 "quantity":"",
 "process":"",
 "failure_category":"",
 "failure_detail":"",
 "issue_summary":"",
 "summary_ko":"",
 "summary_zh":"",
 "summary_en":"",
 "spec":"",
 "actual":"",
 "judgement":"",
 "occurrence_condition":"",
 "defect_rate":"",
 "customer_impact":"",
 "priority":"",
 "customer_request":[""],
 "response_due":"",
 "faca_due":"",
 "internal_action_items":[""],
 "ai_suggested_cause":[""],
 "required_check_points":[""],
 "missing_information":[""]
}}

Allowed process:
{PROCESSES}

Allowed failure_category:
{FAILURES}

Allowed priority:
{PRIORITIES}

VOC SOURCE:
{text}
"""

def call_openai(prompt, api_key, model):
    if OpenAI is None:
        raise RuntimeError("openai 패키지가 설치되지 않았습니다.")
    client = OpenAI(api_key=api_key)
    resp = client.responses.create(model=model, input=prompt)
    return resp.output_text

def call_gemini(prompt, api_key, model):
    if genai is None:
        raise RuntimeError("google-genai 패키지가 설치되지 않았습니다.")
    client = genai.Client(api_key=api_key)
    resp = client.models.generate_content(model=model, contents=prompt)
    return resp.text

def call_openrouter(prompt, api_key, model):
    payload = json.dumps({
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost:8501",
            "X-Title": "VOC Intelligence"
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as res:
            data = json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"OpenRouter API 오류 {e.code}: {body[:500]}")
    return data["choices"][0]["message"]["content"]

def call_provider(provider, prompt, api_key, model):
    if provider == "OpenAI":
        return call_openai(prompt, api_key, model)
    if provider == "Google Gemini":
        return call_gemini(prompt, api_key, model)
    if provider == "OpenRouter Free":
        return call_openrouter(prompt, api_key, model)
    raise RuntimeError("지원하지 않는 AI Provider입니다.")

def is_transient_api_error(exc):
    msg = str(exc).lower()
    tokens = [
        "503", "502", "504", "500", "429", "unavailable", "high demand",
        "overloaded", "rate limit", "resource_exhausted", "timeout",
        "timed out", "temporarily"
    ]
    return any(t in msg for t in tokens)

def _contains_hangul(text):
    return bool(re.search(r"[가-힣]", str(text or "")))

def ensure_multilingual_summaries(result, source_text, provider, api_key, model):
    """Repair missing multilingual summaries and force issue_summary to Korean.

    Most providers follow the main JSON prompt, but some models occasionally omit
    summary fields or mirror the source language into issue_summary.  Only when
    that happens, run a small repair request instead of paying for a second full
    analysis on every VOC.
    """
    result = dict(result or {})
    needs_repair = (
        not str(result.get("summary_ko", "")).strip()
        or not str(result.get("summary_zh", "")).strip()
        or not str(result.get("summary_en", "")).strip()
        or not _contains_hangul(result.get("issue_summary", ""))
    )
    if not needs_repair:
        return result

    repair_prompt = f"""
You are repairing only the language-summary fields of an NPI/VOC analysis.
Use the VOC source and the existing extracted analysis below.

MANDATORY OUTPUT RULES:
- Return ONLY valid JSON with exactly four keys.
- issue_summary: concise professional KOREAN only.
- summary_ko: concise professional KOREAN summary.
- summary_zh: concise professional SIMPLIFIED CHINESE summary.
- summary_en: concise professional ENGLISH summary.
- None of the four values may be blank when meaningful VOC information exists.
- Preserve product codes, LOT numbers, dates, quantities, units, SPEC and Actual values exactly.
- Do not invent facts. Root-cause hypotheses must not be presented as confirmed facts.

Return format:
{{"issue_summary":"","summary_ko":"","summary_zh":"","summary_en":""}}

VOC SOURCE:
{source_text}

EXISTING ANALYSIS:
{json.dumps(result, ensure_ascii=False)}
"""
    try:
        raw = call_provider(provider, repair_prompt, api_key, model)
        repaired = json.loads(clean_json(raw))
        for key in ["issue_summary", "summary_ko", "summary_zh", "summary_en"]:
            value = str(repaired.get(key, "") or "").strip()
            if value:
                result[key] = value
    except Exception:
        # Keep the successful main analysis even if the small repair request fails.
        pass

    # Final safe fallback: if Korean summary exists, issue_summary should never
    # remain in the source language.
    if not _contains_hangul(result.get("issue_summary", "")) and str(result.get("summary_ko", "")).strip():
        result["issue_summary"] = result["summary_ko"]
    return result

def deep_analyze(text, provider, api_key, model):
    if provider == "Local / No API" or not api_key:
        return heuristic_analysis(text)

    prompt = build_analysis_prompt(text)
    last_error = None
    for attempt in range(2):
        try:
            raw = call_provider(provider, prompt, api_key, model)
            result = json.loads(clean_json(raw))
            result = ensure_multilingual_summaries(result, text, provider, api_key, model)
            result["analysis_source"] = f"{provider} ({model})"
            if attempt:
                result["analysis_warning"] = "AI 서버 혼잡으로 재시도 후 분석에 성공했습니다."
            return result
        except Exception as e:
            last_error = e
            if not is_transient_api_error(e):
                raise
            if attempt == 0:
                time.sleep(2)

    if last_error and is_transient_api_error(last_error):
        result = heuristic_analysis(text)
        result["analysis_source"] = "Local / Rule-based · Emergency Fallback"
        result["analysis_warning"] = "AI 서버가 일시적으로 혼잡하여 규칙 기반 결과를 표시합니다."
        return result
    raise last_error

def run_analysis_with_progress(text, provider, api_key, model):
    """Run the blocking AI request in a worker thread while the UI shows estimated progress.

    Gemini/OpenAI do not expose true token-by-token job completion for this request,
    so the percentage is intentionally labeled as an estimate. It never reaches
    100% until the actual analysis has completed.
    """
    progress = st.progress(0, text="분석 준비 중 · Preparing analysis · 准备分析")
    status = st.empty()
    started = time.time()

    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(deep_analyze, text, provider, api_key, model)
        while not future.done():
            elapsed = int(time.time() - started)

            if elapsed < 3:
                pct, stage = 8, "VOC 원문과 분석 조건을 확인하고 있습니다. / Checking input"
            elif elapsed < 10:
                pct, stage = 18, "AI 서버에 분석 요청을 전송하고 있습니다. / Sending request"
            elif elapsed < 25:
                pct, stage = 30, "핵심 정보와 SPEC/Actual을 추출하고 있습니다. / Extracting facts"
            elif elapsed < 45:
                pct, stage = 42, "VOC 원인 가설과 확인 항목을 분석하고 있습니다. / Analyzing checkpoints"
            elif elapsed < 75:
                pct, stage = 55, "AI 응답을 기다리고 있습니다. / Waiting for AI response"
            elif elapsed < 110:
                pct, stage = 66, "심층 분석을 계속 진행하고 있습니다. / Deep analysis in progress"
            elif elapsed < 150:
                pct, stage = 75, "결과 구조화 및 업무 항목 정리 중입니다. / Structuring results"
            elif elapsed < 210:
                pct, stage = 84, "한·중·영 요약을 확인하고 있습니다. / Checking multilingual summaries"
            else:
                pct = min(95, 84 + ((elapsed - 210) // 30 + 1) * 2)
                stage = "AI 응답이 평소보다 오래 걸리고 있습니다. 계속 처리 중입니다. / Still processing"

            progress.progress(
                int(pct),
                text=f"예상 진행률 {int(pct)}% · Estimated progress · 已用 {elapsed}초"
            )
            status.info(f"⏳ {stage}  ·  경과시간 {elapsed}초")
            time.sleep(1)

        result = future.result()

    elapsed = int(time.time() - started)
    progress.progress(100, text=f"분석 완료 100% · Completed · 총 {elapsed}초")
    status.success(f"✅ AI 분석이 완료되었습니다. 총 소요시간: {elapsed}초")
    time.sleep(0.4)
    return result


def translate_with_provider(text, provider, api_key, model):
    if provider == "Local / No API" or not api_key:
        raise RuntimeError("한·중·영 변환은 AI Provider와 API Key가 필요합니다.")
    prompt = f"""Translate the following NPI/VOC technical text into:
1) Korean
2) Simplified Chinese
3) concise professional English.
Preserve product codes, LOT, dates, quantities, units and technical meaning.
Return ONLY valid JSON: {{"ko":"","zh":"","en":""}}
TEXT:
{text}"""
    raw = call_provider(provider, prompt, api_key, model)
    return json.loads(clean_json(raw))

# ---------- UI helpers ----------
def due_bucket(row):
    if str(row.get("status", "")) in ("Closed", "Drop"):
        return None
    dates = []
    for col in ["response_due", "faca_due"]:
        v = row.get(col)
        if pd.notna(v) and str(v).strip():
            d = pd.to_datetime(v, errors="coerce")
            if pd.notna(d):
                dates.append(d.date())
    if not dates:
        return None
    delta = (min(dates) - date.today()).days
    if delta < 0:
        return "Overdue"
    if delta == 0:
        return "Today"
    if delta <= 3:
        return "≤3 Days"
    return None

def require_login():
    if "user" not in st.session_state:
        st.session_state.user = None
    if st.session_state.user:
        return st.session_state.user

    st.markdown("""
    <div style="max-width:540px;margin:3.5rem auto 1.2rem auto;text-align:center;">
      <div style="font-family:Georgia,'Times New Roman',serif;letter-spacing:.14em;
                  font-size:1.8rem;font-weight:700;color:#9A742D;">VOC INTELLIGENCE</div>
      <div style="margin-top:.35rem;color:#6B7280;font-size:.9rem;">
        Korea HQ ↔ China Entity · Shared VOC Workspace
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.caption(APP_VERSION)
    lang_name = st.selectbox(
        "Language / 언어 / 语言",
        list(LANG_OPTIONS.keys()),
        index=0 if st.session_state.get("lang", "ko") == "ko" else 1,
    )
    st.session_state.lang = LANG_OPTIONS[lang_name]
    st.markdown(tr("subtitle"))

    with st.form("login_form"):
        username = st.text_input(tr("login_id"))
        password = st.text_input(tr("login_pw"), type="password")
        submitted = st.form_submit_button(tr("login"))
        if submitted:
            user = authenticate(username, password)
            if user:
                st.session_state.user = user
                st.rerun()
            else:
                st.error("로그인 정보가 올바르지 않습니다. / 登录信息不正确。")
    st.info("최초 관리자 / Initial Admin: admin / ChangeMe123!")
    st.stop()


def render_top_header():
    st.markdown("""
    <style>
      :root {
        --voc-bg:#F6F7F9;
        --voc-card:#FFFFFF;
        --voc-text:#20232A;
        --voc-muted:#6B7280;
        --voc-line:#E7E9EE;
        --voc-gold:#B58A3A;
        --voc-gold-soft:#F7F1E5;
      }
      .stApp { background:var(--voc-bg); color:var(--voc-text); }
      .block-container { padding-top:.75rem; padding-bottom:2.5rem; max-width:1440px; }
      [data-testid="stSidebar"], section[data-testid="stSidebar"] { display:none; }
      header[data-testid="stHeader"] { background:transparent; }

      .voc-shell {
        background:var(--voc-card);
        border:1px solid var(--voc-line);
        border-radius:18px;
        padding:14px 18px;
        margin-bottom:14px;
        box-shadow:0 5px 22px rgba(24,28,36,.045);
      }
      .voc-brand-row {
        display:flex; align-items:center; justify-content:space-between; gap:18px;
      }
      .voc-brand-title {
        font-family:Georgia,'Times New Roman',serif;
        letter-spacing:.13em;
        font-size:1.28rem;
        font-weight:700;
        color:#9A742D;
      }
      .voc-version { font-size:.74rem; color:var(--voc-muted); margin-top:3px; }
      .voc-chip {
        display:inline-flex; align-items:center; gap:6px;
        background:#F8F4EA; color:#80601F;
        border:1px solid #E9DDBF;
        border-radius:999px;
        padding:5px 10px;
        font-size:.74rem;
        font-weight:650;
        white-space:nowrap;
      }
      .voc-section-note {
        color:var(--voc-muted);
        font-size:.88rem;
        margin-top:-.45rem;
        margin-bottom:1rem;
      }
      div[data-testid="stMetric"] {
        border:1px solid var(--voc-line);
        padding:16px;
        border-radius:16px;
        background:#fff;
        box-shadow:0 3px 14px rgba(24,28,36,.035);
      }
      div[data-testid="stMetric"] label { color:var(--voc-muted)!important; }
      div[data-testid="stMetricValue"] { font-weight:700; }

      .stButton > button, .stDownloadButton > button {
        border-radius:11px;
        min-height:2.75rem;
      }
      .stTextInput input, .stTextArea textarea, div[data-baseweb="select"] > div {
        border-radius:10px!important;
      }

      div[data-baseweb="tab-list"] {
        gap:.55rem;
        border:0!important;
        background:#fff;
        padding:7px;
        border-radius:15px;
        box-shadow:0 3px 14px rgba(24,28,36,.04);
        margin-bottom:1.2rem;
      }
      button[data-baseweb="tab"] {
        min-width:62px;
        height:48px;
        border-radius:11px!important;
        font-size:1.28rem!important;
        padding:.4rem .8rem!important;
      }
      button[data-baseweb="tab"][aria-selected="true"] {
        background:var(--voc-gold-soft)!important;
        color:#8A6827!important;
      }
      button[data-baseweb="tab"]:hover { background:#F4F5F7!important; }

      div[data-testid="stVerticalBlockBorderWrapper"] {
        border-color:var(--voc-line)!important;
        border-radius:16px!important;
      }
      h1,h2,h3 { letter-spacing:-.02em; }
      hr { border-color:var(--voc-line); }


      /* Reliable icon navigation buttons */
      .voc-nav-wrap {
        background:#FFFFFF;
        border:1px solid var(--voc-line);
        border-radius:15px;
        padding:7px;
        margin-bottom:1.1rem;
        box-shadow:0 3px 14px rgba(24,28,36,.04);
      }

      .st-key-nav_dashboard button,
      .st-key-nav_analysis button,
      .st-key-nav_search button,
      .st-key-nav_translate button,
      .st-key-nav_settings button {
        width:100%;
        height:46px;
        min-height:46px;
        border:0!important;
        border-radius:10px!important;
        background:transparent!important;
        font-size:1.18rem!important;
        transition:background .18s ease, color .18s ease!important;
      }

      .st-key-nav_dashboard button:hover,
      .st-key-nav_analysis button:hover,
      .st-key-nav_search button:hover,
      .st-key-nav_translate button:hover,
      .st-key-nav_settings button:hover {
        background:#F7F1E5!important;
        color:#8A6827!important;
      }

      .st-key-nav_dashboard button:hover p,
      .st-key-nav_analysis button:hover p,
      .st-key-nav_search button:hover p,
      .st-key-nav_translate button:hover p,
      .st-key-nav_settings button:hover p {
        font-size:0!important;
      }

      .st-key-nav_dashboard button:hover p::after {
        content:"Dashboard";
        font-size:.83rem;
        font-weight:700;
      }
      .st-key-nav_analysis button:hover p::after {
        content:"AI 분석";
        font-size:.83rem;
        font-weight:700;
      }
      .st-key-nav_search button:hover p::after {
        content:"VOC 검색";
        font-size:.83rem;
        font-weight:700;
      }
      .st-key-nav_translate button:hover p::after {
        content:"Translate";
        font-size:.83rem;
        font-weight:700;
      }
      .st-key-nav_settings button:hover p::after {
        content:"Settings";
        font-size:.83rem;
        font-weight:700;
      }

      @media (max-width:800px) {
        .block-container { padding-left:.75rem; padding-right:.75rem; }
        button[data-baseweb="tab"] { min-width:48px; padding:.3rem .55rem!important; }
        .voc-brand-title { font-size:1.05rem; }
        .voc-chip { display:none; }
      }
    </style>
    """, unsafe_allow_html=True)

    backend_short = "Shared DB" if DATABASE_URL else "Local DB"
    st.markdown(
        f"""
        <div class="voc-shell">
          <div class="voc-brand-row">
            <div>
              <div class="voc-brand-title">VOC INTELLIGENCE</div>
              <div class="voc-version">{APP_VERSION} · Korea HQ ↔ China Entity</div>
            </div>
            <div class="voc-chip">● {backend_short}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )

def dashboard(df):
    st.header("Overview")
    st.markdown(
        '<div class="voc-section-note">VOC 현황과 우선 대응 항목을 한눈에 확인합니다.</div>',
        unsafe_allow_html=True
    )

    if df.empty:
        st.info("아직 등록된 VOC가 없습니다. / 暂无VOC数据。")
        return

    statuses = df["status"].fillna("")
    tmp = df.copy()
    tmp["Due"] = tmp.apply(due_bucket, axis=1)
    action = tmp[tmp["Due"].notna()]

    metrics = [
        ("Total VOC", len(df)),
        ("Active", int((~statuses.isin(["Closed", "Drop"])).sum())),
        ("Pending", int((statuses == "Pending").sum())),
        ("Due / Overdue", len(action)),
        ("Closed", int((statuses == "Closed").sum())),
    ]
    cols = st.columns(5)
    for c, (label, value) in zip(cols, metrics):
        c.metric(label, value)

    st.subheader("Priority Actions")
    if action.empty:
        st.success("현재 긴급 Due 항목이 없습니다. / No urgent due items / 暂无紧急到期项目")
    else:
        st.dataframe(
            action[
                ["voc_id", "customer", "product", "status",
                 "response_due", "faca_due", "Due", "dri"]
            ],
            use_container_width=True,
            hide_index=True
        )

    st.subheader("Recent VOC")
    recent_cols = [
        "voc_id", "received_date", "customer", "product",
        "failure_category", "priority", "status", "dri"
    ]
    st.dataframe(df[recent_cols].head(8), use_container_width=True, hide_index=True)

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            st.markdown("#### Customer")
            st.bar_chart(df["customer"].replace("", "(Blank)").value_counts().head(8))
    with c2:
        with st.container(border=True):
            st.markdown("#### Failure Category")
            st.bar_chart(df["failure_category"].replace("", "(Blank)").value_counts().head(8))

def ai_analysis_page(user, provider, api_key, model):
    st.header("AI VOC Analysis")
    st.markdown(
        '<div class="voc-section-note">고객 VOC를 구조화하고, NPI 대응에 필요한 핵심 정보와 가설을 정리합니다.</div>',
        unsafe_allow_html=True
    )
    st.caption("고객 메일·WeChat·VOC 원문을 그대로 붙여넣으면 AI가 항목별로 구조화합니다.")

    original = st.text_area(
        UI["original"], key="ai_original", height=220,
        placeholder="고객 메일 / WeChat / VOC 내용을 그대로 붙여넣으세요."
    )

    if st.button("🧠 VOC 심층 분석 / Deep Analyze / 深度分析", type="primary", use_container_width=True):
        if not original.strip():
            st.warning("VOC 원문을 입력하세요.")
        else:
            try:
                r = run_analysis_with_progress(original, provider, api_key, model)
                mappings = [
                    "product","material_type","lot_no","quantity","process",
                    "failure_category","failure_detail","issue_summary",
                    "summary_ko","summary_zh","summary_en","spec","actual",
                    "judgement","occurrence_condition","defect_rate",
                    "customer_impact","priority","response_due","faca_due"
                ]
                for k in mappings:
                    val = r.get(k, "")
                    if k == "process" and val not in PROCESSES:
                        val = ""
                    if k == "failure_category" and val not in FAILURES:
                        val = ""
                    if k == "priority" and val not in PRIORITIES:
                        val = "High"
                    st.session_state[f"ai_{k}"] = val
                for k in [
                    "customer_request","internal_action_items","ai_suggested_cause",
                    "required_check_points","missing_information"
                ]:
                    st.session_state[f"ai_{k}"] = list_to_text(
                        r.get(k, []), numbered=(k == "customer_request")
                    )
                st.session_state["analysis_source"] = r.get("analysis_source","")
                st.session_state["analysis_warning"] = r.get("analysis_warning","")
            except Exception as e:
                st.error(f"AI 분석 오류: {e}")

    if st.session_state.get("analysis_source"):
        st.success(f"분석 완료 · {st.session_state['analysis_source']}")
        if st.session_state.get("analysis_warning"):
            st.warning(st.session_state["analysis_warning"])

    st.subheader("1. 핵심 추출 결과 / Core Extraction / 核心提取")
    a,b,c = st.columns(3)
    with a:
        received = st.date_input(UI["date"], value=date.today(), key="ai_received_date")
        customer = st.text_input(UI["customer"], key="ai_customer")
        dri = st.text_input(UI["dri"], value=user["display_name"], key="ai_dri")
        product = st.text_input(UI["product"], key="ai_product")
        material = st.text_input(UI["material"], key="ai_material_type")
    with b:
        structure = st.text_input(UI["structure"], key="ai_product_structure")
        lot = st.text_input(UI["lot"], key="ai_lot_no")
        qty = st.text_input(UI["quantity"], key="ai_quantity")
        process = st.selectbox(
            UI["process"], PROCESSES,
            index=PROCESSES.index(st.session_state.get("ai_process","")) if st.session_state.get("ai_process","") in PROCESSES else 0,
            format_func=lambda x: PROCESS_DISPLAY.get(x,x), key="ai_process_widget"
        )
        failure = st.selectbox(
            UI["failure"], FAILURES,
            index=FAILURES.index(st.session_state.get("ai_failure_category","")) if st.session_state.get("ai_failure_category","") in FAILURES else 0,
            format_func=lambda x: FAILURE_DISPLAY.get(x,x), key="ai_failure_widget"
        )
    with c:
        detail = st.text_input(UI["detail"], key="ai_failure_detail")
        spec = st.text_input(UI["spec"], key="ai_spec")
        actual = st.text_input(UI["actual"], key="ai_actual")
        judgement = st.text_input(UI["judgement"], key="ai_judgement")
        priority_default = st.session_state.get("ai_priority","High")
        priority = st.selectbox(
            UI["priority"], PRIORITIES,
            index=PRIORITIES.index(priority_default) if priority_default in PRIORITIES else 2,
            key="ai_priority_widget"
        )

    issue_summary = st.text_area(UI["summary"], key="ai_issue_summary", height=90)
    occurrence = st.text_input(UI["condition"], key="ai_occurrence_condition")
    defect_rate = st.text_input(UI["rate"], key="ai_defect_rate")
    impact = st.text_area(UI["impact"], key="ai_customer_impact", height=85)

    st.subheader("2. 고객 요구 & Deadline / Customer Request / 客户要求")
    x,y,z = st.columns([2,1,1])
    with x:
        request = st.text_area(UI["request"], key="ai_customer_request", height=150)
    with y:
        response_due = st.text_input(UI["response_due"], key="ai_response_due")
    with z:
        faca_due = st.text_input(UI["faca_due"], key="ai_faca_due")

    st.subheader("3. NPI 대응 Assistant")
    p,q = st.columns(2)
    with p:
        internal = st.text_area("🚨 " + UI["internal"], key="ai_internal_action_items", height=180)
        checks = st.text_area("🔍 " + UI["check"], key="ai_required_check_points", height=180)
    with q:
        causes = st.text_area("💡 " + UI["cause"], key="ai_ai_suggested_cause", height=180)
        missing = st.text_area("❓ " + UI["missing"], key="ai_missing_information", height=180)

    st.subheader("4. 한 · 중 · 영 요약 / KO · ZH · EN Summary")
    l1,l2,l3 = st.columns(3)
    with l1:
        s_ko = st.text_area(UI["ko"], key="ai_summary_ko", height=150)
    with l2:
        s_zh = st.text_area(UI["zh"], key="ai_summary_zh", height=150)
    with l3:
        s_en = st.text_area(UI["en"], key="ai_summary_en", height=150)

    with st.expander("5. FACA / 확정 분석 / Confirmed Analysis / 确认分析"):
        root = st.text_area(UI["root"], key="ai_confirmed_root")
        escape = st.text_area(UI["escape"], key="ai_escape_cause")
        corrective = st.text_area(UI["corrective"], key="ai_corrective_action")
        verification = st.text_area(UI["verification"], key="ai_verification_result")
        faca_no = st.text_input(UI["faca_no"], key="ai_faca_no")
        status = st.selectbox(UI["status"], STATUS_OPTIONS, key="ai_status")
        picture = st.text_input(UI["picture"], key="ai_picture_link")
        remark = st.text_area(UI["remark"], key="ai_remark")

    if user["role"] != "Viewer":
        if st.button("💾 검토 완료 → 중앙 DB 저장 / Save to Central DB / 保存到中央数据库", use_container_width=True):
            data = {
                "voc_id": next_voc_id(),
                "received_date": str(received),
                "customer": customer, "dri": dri, "product": product,
                "material_type": material, "product_structure": structure,
                "lot_no": lot, "quantity": qty, "process": process,
                "failure_category": failure, "failure_detail": detail,
                "voc_original_text": original, "issue_summary": issue_summary,
                "summary_ko": s_ko, "summary_zh": s_zh, "summary_en": s_en,
                "spec": spec, "actual": actual, "judgement": judgement,
                "occurrence_condition": occurrence, "defect_rate": defect_rate,
                "customer_impact": impact, "priority": priority,
                "customer_request": request, "response_due": response_due,
                "faca_due": faca_due, "internal_action_items": internal,
                "ai_suggested_cause": causes, "required_check_points": checks,
                "missing_information": missing, "confirmed_root_cause": root,
                "escape_cause": escape, "corrective_action": corrective,
                "verification_result": verification, "faca_no": faca_no,
                "status": status, "close_date": "", "picture_link": picture,
                "remark": remark,
            }
            try:
                vid = save_case(data, user["username"])
                st.success(f"✅ {vid} 중앙 DB 저장 완료")
            except Exception as e:
                st.error(f"저장 실패: {e}")

def search_edit_page(user, df):
    st.header("VOC Search & Edit")
    st.markdown(
        '<div class="voc-section-note">VOC를 검색하고 수정합니다. Admin은 목록 오른쪽 × 버튼으로 잘못 등록된 VOC를 삭제할 수 있습니다.</div>',
        unsafe_allow_html=True
    )

    if df.empty:
        st.info("저장된 VOC가 없습니다. / 暂无已保存VOC。")
        return

    q = st.text_input(
        "VOC 검색 / Search / 搜索",
        placeholder="VOC ID / Product / LOT / Customer"
    )

    show = df.copy()
    if q:
        mask = pd.Series(False, index=show.index)
        for col in [
            "voc_id", "product", "lot_no", "customer",
            "failure_category", "issue_summary"
        ]:
            mask |= show[col].fillna("").astype(str).str.contains(
                q, case=False, na=False
            )
        show = show[mask]

    if show.empty:
        st.warning("검색 결과가 없습니다.")
        return

    # Compact row list. Admin sees an X button at the far right.
    header_cols = st.columns([1.35, 1.25, 1.5, 1.35, 1.1, .9, .55])
    headers = ["VOC ID", "Date", "Customer", "Product", "Status", "DRI", ""]
    for col, label in zip(header_cols, headers):
        col.markdown(f"**{label}**")

    for _, row in show.head(100).iterrows():
        voc_id = str(row.get("voc_id", "") or "")
        cols = st.columns([1.35, 1.25, 1.5, 1.35, 1.1, .9, .55])
        cols[0].write(voc_id)
        cols[1].write(str(row.get("received_date", "") or ""))
        cols[2].write(str(row.get("customer", "") or ""))
        cols[3].write(str(row.get("product", "") or ""))
        cols[4].write(str(row.get("status", "") or ""))
        cols[5].write(str(row.get("dri", "") or ""))

        if user["role"] == "Admin":
            if cols[6].button(
                "×",
                key=f"row_delete_{voc_id}",
                help=f"{voc_id} 삭제",
                use_container_width=True
            ):
                st.session_state["pending_delete_voc"] = voc_id
        else:
            cols[6].write("")

    pending_delete = st.session_state.get("pending_delete_voc")
    if pending_delete:
        st.error(
            f"⚠️ **{pending_delete}** VOC를 삭제하시겠습니까?\n\n"
            "삭제하면 해당 VOC와 변경 이력(Audit Log)이 중앙 DB에서 영구 삭제되며 복구할 수 없습니다."
        )
        dc1, dc2, dc3 = st.columns([1, 1, 4])
        with dc1:
            if st.button(
                "삭제",
                type="primary",
                key=f"confirm_delete_{pending_delete}",
                use_container_width=True
            ):
                try:
                    delete_case(pending_delete, user["username"])
                    st.session_state.pop("pending_delete_voc", None)
                    st.success(f"✅ {pending_delete} 삭제 완료")
                    st.rerun()
                except Exception as e:
                    st.error(f"삭제 실패: {e}")
        with dc2:
            if st.button(
                "취소",
                key=f"cancel_delete_{pending_delete}",
                use_container_width=True
            ):
                st.session_state.pop("pending_delete_voc", None)
                st.rerun()

    st.divider()

    ids = show["voc_id"].astype(str).tolist()
    selected = st.selectbox(
        "상세보기 / 수정할 VOC / Select VOC / 选择VOC",
        ids,
        key="search_selected_voc"
    )
    rec = load_case(selected)
    if not rec:
        return

    if user["role"] == "Viewer":
        st.json(rec)
        return

    e1, e2, e3 = st.columns(3)
    with e1:
        customer = st.text_input(
            UI["customer"], value=rec.get("customer") or "", key="ed_customer"
        )
        dri = st.text_input(
            UI["dri"], value=rec.get("dri") or "", key="ed_dri"
        )
        product = st.text_input(
            UI["product"], value=rec.get("product") or "", key="ed_product"
        )
        lot = st.text_input(
            UI["lot"], value=rec.get("lot_no") or "", key="ed_lot"
        )

    with e2:
        process_val = rec.get("process") or ""
        process = st.selectbox(
            UI["process"], PROCESSES,
            index=PROCESSES.index(process_val) if process_val in PROCESSES else 0,
            format_func=lambda x: PROCESS_DISPLAY.get(x, x),
            key="ed_process"
        )

        failure_val = rec.get("failure_category") or ""
        failure = st.selectbox(
            UI["failure"], FAILURES,
            index=FAILURES.index(failure_val) if failure_val in FAILURES else 0,
            format_func=lambda x: FAILURE_DISPLAY.get(x, x),
            key="ed_failure"
        )

        priority_val = rec.get("priority") or "High"
        priority = st.selectbox(
            UI["priority"], PRIORITIES,
            index=PRIORITIES.index(priority_val) if priority_val in PRIORITIES else 2,
            key="ed_priority"
        )

        status_val = rec.get("status") or "Open"
        status = st.selectbox(
            UI["status"], STATUS_OPTIONS,
            index=STATUS_OPTIONS.index(status_val) if status_val in STATUS_OPTIONS else 0,
            key="ed_status"
        )

    with e3:
        response_due = st.text_input(
            UI["response_due"],
            value=rec.get("response_due") or "",
            key="ed_response_due"
        )
        faca_due = st.text_input(
            UI["faca_due"],
            value=rec.get("faca_due") or "",
            key="ed_faca_due"
        )
        faca_no = st.text_input(
            UI["faca_no"],
            value=rec.get("faca_no") or "",
            key="ed_faca_no"
        )

    summary = st.text_area(
        UI["summary"], value=rec.get("issue_summary") or "", key="ed_summary"
    )
    request = st.text_area(
        UI["request"], value=rec.get("customer_request") or "", key="ed_request"
    )
    actions = st.text_area(
        UI["internal"], value=rec.get("internal_action_items") or "", key="ed_actions"
    )
    checks = st.text_area(
        UI["check"], value=rec.get("required_check_points") or "", key="ed_checks"
    )
    root = st.text_area(
        UI["root"], value=rec.get("confirmed_root_cause") or "", key="ed_root"
    )
    escape = st.text_area(
        UI["escape"], value=rec.get("escape_cause") or "", key="ed_escape"
    )
    corrective = st.text_area(
        UI["corrective"], value=rec.get("corrective_action") or "", key="ed_corrective"
    )
    verification = st.text_area(
        UI["verification"],
        value=rec.get("verification_result") or "",
        key="ed_verification"
    )

    update_summary = st.text_input(
        "이번 수정 내용 / Update Summary / 本次修改内容",
        placeholder="예: 불량 Sample 확보 및 FACA 원인분석 결과 업데이트"
    )

    if st.button(
        "💾 수정 내용 저장 / Save Changes / 保存修改",
        type="primary",
        use_container_width=True
    ):
        if not update_summary.strip():
            st.warning("수정 이력을 입력해주세요.")
        else:
            update_case(
                selected,
                {
                    "customer": customer,
                    "dri": dri,
                    "product": product,
                    "lot_no": lot,
                    "process": process,
                    "failure_category": failure,
                    "priority": priority,
                    "status": status,
                    "response_due": response_due,
                    "faca_due": faca_due,
                    "faca_no": faca_no,
                    "issue_summary": summary,
                    "customer_request": request,
                    "internal_action_items": actions,
                    "required_check_points": checks,
                    "confirmed_root_cause": root,
                    "escape_cause": escape,
                    "corrective_action": corrective,
                    "verification_result": verification,
                },
                user["username"],
                update_summary.strip()
            )
            st.success("✅ 수정 완료")

    st.subheader("변경 이력 / Audit Log / 变更记录")
    st.dataframe(load_audit(selected), use_container_width=True, hide_index=True)

def translation_page(provider, api_key, model):
    st.header("Translate")
    src = st.text_area("업무 문장 / Business Text / 工作文本", height=180)
    if st.button("🌐 변환 / Translate / 翻译"):
        if not src.strip():
            st.warning("텍스트를 입력하세요.")
        else:
            try:
                r = translate_with_provider(src, provider, api_key, model)
                c1,c2,c3 = st.columns(3)
                with c1:
                    st.text_area("한국어 / Korean / 韩文", r.get("ko",""), height=220)
                with c2:
                    st.text_area("中文（简体） / Chinese", r.get("zh",""), height=220)
                with c3:
                    st.text_area("English / 영어 / 英文", r.get("en",""), height=220)
            except Exception as e:
                st.error(str(e))

def settings_page(user):
    st.header("Settings")
    st.caption(f"Database Backend: {DB_BACKEND}")
    if DATABASE_URL:
        st.success("☁️ 중앙 PostgreSQL DB 연결됨 / Shared DB Connected / 已连接共享数据库")
    else:
        st.warning("💻 현재 Local SQLite 모드입니다. 공개 웹앱 배포 시 DATABASE_URL을 Secrets에 설정하세요.")

    st.subheader("내 비밀번호 변경 / Change Password / 修改密码")
    with st.form("change_pw"):
        p1 = st.text_input("새 비밀번호 / New Password / 新密码", type="password")
        p2 = st.text_input("비밀번호 확인 / Confirm / 确认密码", type="password")
        if st.form_submit_button("비밀번호 변경 / Change / 修改"):
            if len(p1) < 8:
                st.error("비밀번호는 8자 이상으로 설정하세요.")
            elif p1 != p2:
                st.error("비밀번호가 일치하지 않습니다.")
            else:
                change_password(user["username"], p1)
                st.success("비밀번호를 변경했습니다.")

    st.subheader("웹 배포 상태 / Cloud Deployment / 云端部署")
    cdb, cai = st.columns(2)
    with cdb:
        st.metric("Database", "Shared PostgreSQL" if DATABASE_URL else "Local SQLite")
    with cai:
        configured = []
        if _secret("GEMINI_API_KEY", ""): configured.append("Gemini")
        if _secret("OPENROUTER_API_KEY", ""): configured.append("OpenRouter")
        if _secret("OPENAI_API_KEY", ""): configured.append("OpenAI")
        st.metric("Server AI Secret", ", ".join(configured) if configured else "Not configured")
    st.caption("공개 테스트 배포에서는 실제 고객/LOT/기밀 자료 대신 더미 또는 비식별 데이터를 사용하세요.")

    if user["role"] in ("Admin", "Editor"):
        st.subheader("📦 중앙 DB 백업 / Database Backup / 数据库备份")
        st.caption(
            "Admin과 Editor는 중앙 DB의 VOC 전체 데이터와 변경 이력을 Excel로 내려받을 수 있습니다. "
            "사용자 비밀번호 정보는 포함하지 않습니다."
        )
        try:
            backup_bytes = build_excel_backup()
            backup_name = f"VOC_DB_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                "⬇️ 전체 DB Excel 백업 다운로드 / Download Full DB Backup / 下载完整数据库备份",
                data=backup_bytes,
                file_name=backup_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"백업 파일 생성 실패: {e}")

    if user["role"] == "Admin":
        st.subheader("사용자 관리 / User Management / 用户管理")
        users = list_users_df()
        st.dataframe(users, use_container_width=True, hide_index=True)
        with st.expander("새 사용자 추가 / Add User / 新增用户"):
            with st.form("add_user"):
                c1,c2 = st.columns(2)
                username = c1.text_input("Username")
                display = c2.text_input("Display Name")
                c3,c4 = st.columns(2)
                role = c3.selectbox("Role", ROLE_OPTIONS, index=1)
                password = c4.text_input("Temporary Password", type="password")
                if st.form_submit_button("사용자 생성 / Create User / 创建用户"):
                    if not username or not password:
                        st.error("Username과 Password가 필요합니다.")
                    else:
                        with ENGINE.begin() as conn:
                            create_user(conn, username, display or username, password, role)
                        st.success("사용자를 생성했습니다.")

        with st.expander("사용자 비밀번호 초기화 / Reset User Password / 重置用户密码"):
            usernames = users["username"].tolist() if not users.empty else []
            if not usernames:
                st.info("등록된 사용자가 없습니다.")
            else:
                with st.form("admin_reset_pw"):
                    target_user = st.selectbox("사용자 선택 / User / 用户", usernames)
                    temp_pw = st.text_input("새 임시 비밀번호 / New Temporary Password / 新临时密码", type="password")
                    temp_pw2 = st.text_input("임시 비밀번호 확인 / Confirm / 确认", type="password")
                    if st.form_submit_button("비밀번호 초기화 / Reset Password / 重置密码"):
                        if len(temp_pw) < 8:
                            st.error("임시 비밀번호는 8자 이상으로 설정하세요.")
                        elif temp_pw != temp_pw2:
                            st.error("비밀번호가 일치하지 않습니다.")
                        else:
                            change_password(target_user, temp_pw)
                            st.success(f"{target_user} 계정의 비밀번호를 초기화했습니다.")


def main():
    if "lang" not in st.session_state:
        st.session_state.lang = "ko"

    loading = st.empty()
    loading.markdown("""
    <div style="position:fixed;inset:0;z-index:999999;background:#ffffff;display:flex;
                align-items:center;justify-content:center;flex-direction:column;">
      <div style="font-family:Georgia,'Times New Roman',serif;font-size:28px;font-weight:700;
                  letter-spacing:.12em;color:#B89046;margin-bottom:14px;">VOC INTELLIGENCE</div>
      <div style="width:42px;height:42px;border:4px solid #eee;border-top-color:#B89046;
                  border-radius:50%;animation:vocspin 0.9s linear infinite;"></div>
      <div style="margin-top:16px;font-size:15px;color:#666;">로딩 중입니다 · Loading · 加载中...</div>
    </div>
    <style>@keyframes vocspin {to {transform:rotate(360deg);}}</style>
    """, unsafe_allow_html=True)
    init_db()
    loading.empty()
    user = require_login()
    render_top_header()

    # Top control row
    ctl1, ctl2, ctl3, ctl4 = st.columns([1.25, 1.25, 1.8, 1.1])
    with ctl1:
        lang_name = st.selectbox(
            "언어 / Language / 语言",
            list(LANG_OPTIONS.keys()),
            index=0 if st.session_state.get("lang","ko") == "ko" else 1,
            key="top_lang"
        )
        st.session_state.lang = LANG_OPTIONS[lang_name]
    with ctl2:
        provider = st.selectbox(
            "AI Provider",
            ["Google Gemini", "OpenRouter Free", "OpenAI", "Local / No API"],
            key="ai_provider"
        )
    with ctl3:
        server_key = ""
        if provider == "Google Gemini":
            server_key = _secret("GEMINI_API_KEY", "")
            model = _secret("GEMINI_MODEL", "gemini-3.6-flash")
            key_label = "Gemini"
        elif provider == "OpenRouter Free":
            server_key = _secret("OPENROUTER_API_KEY", "")
            model = _secret("OPENROUTER_MODEL", "openrouter/free")
            key_label = "OpenRouter"
        elif provider == "OpenAI":
            server_key = _secret("OPENAI_API_KEY", "")
            model = _secret("OPENAI_MODEL", "gpt-5.6")
            key_label = "OpenAI"
        else:
            model = ""
            key_label = "Local"

        if provider == "Local / No API":
            api_key = ""
            st.caption("Local / Rule-based")
        elif server_key:
            api_key = server_key
            st.success(f"🔐 {key_label} Server Secret 사용")
        else:
            api_key = st.text_input(
                f"{key_label} API Key",
                type="password",
                label_visibility="collapsed",
                placeholder=f"{key_label} API Key · 현재 브라우저 세션에서만 사용"
            )
    with ctl4:
        st.caption(f"{user['display_name']} · {user['role']}")
        if st.button("로그아웃 / Logout / 退出", use_container_width=True):
            st.session_state.user = None
            st.rerun()

    df = load_cases()

    if "current_page" not in st.session_state:
        st.session_state.current_page = "dashboard"

    st.markdown('<div class="voc-nav-wrap">', unsafe_allow_html=True)
    n1, n2, n3, n4, n5 = st.columns(5)

    with n1:
        with st.container(key="nav_dashboard"):
            if st.button(
                "⌂",
                key="nav_btn_dashboard",
                help="Dashboard · 대시보드 · 仪表盘",
                use_container_width=True
            ):
                st.session_state.current_page = "dashboard"

    with n2:
        with st.container(key="nav_analysis"):
            if st.button(
                "✦",
                key="nav_btn_analysis",
                help="AI VOC Analysis · 심층분석 · AI深度分析",
                use_container_width=True
            ):
                st.session_state.current_page = "analysis"

    with n3:
        with st.container(key="nav_search"):
            if st.button(
                "⌕",
                key="nav_btn_search",
                help="Search & Edit · VOC 검색/수정 · 查询修改",
                use_container_width=True
            ):
                st.session_state.current_page = "search"

    with n4:
        with st.container(key="nav_translate"):
            if st.button(
                "文",
                key="nav_btn_translate",
                help="Translate · 한중영 변환 · 韩中英转换",
                use_container_width=True
            ):
                st.session_state.current_page = "translate"

    with n5:
        with st.container(key="nav_settings"):
            if st.button(
                "⚙",
                key="nav_btn_settings",
                help="Settings · 설정 · 设置",
                use_container_width=True
            ):
                st.session_state.current_page = "settings"

    st.markdown('</div>', unsafe_allow_html=True)

    page = st.session_state.current_page
    if page == "dashboard":
        dashboard(df)
    elif page == "analysis":
        ai_analysis_page(user, provider, api_key, model)
    elif page == "search":
        search_edit_page(user, df)
    elif page == "translate":
        translation_page(provider, api_key, model)
    elif page == "settings":
        settings_page(user)

if __name__ == "__main__":
    main()
