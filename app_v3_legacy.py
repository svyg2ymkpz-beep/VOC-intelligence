
import os, re, json, shutil, time
import urllib.request
import urllib.error
from pathlib import Path
from datetime import date, datetime
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    from google import genai
except Exception:
    genai = None

BASE = Path(__file__).resolve().parent
TEMPLATE = BASE / "assets" / "VOC_Master_AI_V2_2.xlsx"
DATA_DIR = BASE / "data"
BACKUP_DIR = BASE / "backups"
DB_PATH = DATA_DIR / "VOC_Master_AI.xlsx"
SHEET = "VOC_Master_AI"

SETTINGS_PATH = DATA_DIR / "settings.json"

DEFAULT_SETTINGS = {
    "current_user": "Jongin",
    "provider": "Google Gemini",
    "gemini_model": "gemini-flash-latest",
    "openrouter_model": "openrouter/free",
    "openai_model": "gpt-5.6-luna",
}

def load_settings():
    DATA_DIR.mkdir(exist_ok=True)
    if not SETTINGS_PATH.exists():
        return DEFAULT_SETTINGS.copy()
    try:
        data = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
        out = DEFAULT_SETTINGS.copy()
        out.update(data or {})
        return out
    except Exception:
        return DEFAULT_SETTINGS.copy()

def save_settings(data):
    DATA_DIR.mkdir(exist_ok=True)
    out = DEFAULT_SETTINGS.copy()
    out.update(data or {})
    SETTINGS_PATH.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

def secret_path(provider):
    safe = re.sub(r"[^A-Za-z0-9]+", "_", provider).strip("_").lower()
    return DATA_DIR / f".{safe}_key"

def load_local_key(provider):
    p = secret_path(provider)
    if p.exists():
        try:
            return p.read_text(encoding="utf-8").strip()
        except Exception:
            return ""
    return ""

def save_local_key(provider, key):
    p = secret_path(provider)
    DATA_DIR.mkdir(exist_ok=True)
    if key and key.strip():
        p.write_text(key.strip(), encoding="utf-8")
        try:
            os.chmod(p, 0o600)
        except Exception:
            pass
    elif p.exists():
        p.unlink()

def due_state(value):
    if value is None or str(value).strip() == "":
        return "none", None
    try:
        d = pd.to_datetime(value, errors="coerce")
        if pd.notna(d):
            dd = d.date()
            delta = (dd - date.today()).days
            if delta < 0:
                return "overdue", dd
            if delta <= 3:
                return "soon", dd
            return "future", dd
    except Exception:
        pass
    return "unknown", None


HEADERS = [
    "VOC ID","Date","Customer","DRI","Product","Material / Product Type","Product Structure","Lot No.",
    "Quantity","Process","Failure Category","Failure Detail","VOC Original Text",
    "Issue Summary","Summary_KO","Summary_ZH","Summary_EN",
    "SPEC","Actual","Judgement","Occurrence Condition","Defect Rate","Customer Impact","Priority",
    "Customer Request","Response Due","FACA Due","Internal Action Items",
    "AI Suggested Cause","Required Check Points","Missing Information",
    "Confirmed Root Cause","Escape Cause","Corrective Action","Verification Result",
    "FACA No.","Status","Close Date","Picture / Link","Remark",
    "Created By","Created Date","Updated By","Updated Date","Last Update Summary"
]

STATUSES = ["Open","In Progress","Pending","Over Due","Waiting Customer","Waiting Internal","Validation","Hold","Drop","Closed"]
PRIORITIES = ["Low","Medium","High","Critical"]
PROCESSES = ["","Customer Incoming / Material","Coating","Drying/Oven","Lamination","Slitting","Die-cut","Waste Stripping","Peeling/Release","Assembly","Reliability Test","Customer Process","Other"]
FAILURES = ["","Appearance","拉胶 / Stringing","溢胶 / Adhesive Overflow","残胶 / Residue","起翘 / Lifting","Curl","Bubble","Wrinkle","Scratch","Foreign Material","Release Force Low","Release Force High","Peel Low","High-temp Peel","Dimension","Other"]

def ensure_db():
    DATA_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)
    if not DB_PATH.exists():
        shutil.copy(TEMPLATE, DB_PATH)

def backup_db():
    ensure_db()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    p = BACKUP_DIR / f"VOC_Master_AI_{stamp}.xlsx"
    shutil.copy(DB_PATH, p)
    return p

def count_records():
    ensure_db()
    wb = load_workbook(DB_PATH, read_only=True, data_only=True)
    ws = wb[SHEET]
    n = sum(1 for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
            if row and row[0] not in (None,""))
    wb.close()
    return n

def next_id(ws):
    year = date.today().year
    mx = 0
    max_row = ws.max_row or 1
    for r in range(2, max_row+1):
        v = ws.cell(r,1).value
        if isinstance(v,str) and v.startswith("VOC-"):
            try: mx = max(mx, int(v.split("-")[-1]))
            except: pass
    return f"VOC-{year}-{mx+1:03d}"

def first_empty_row(ws):
    max_row = ws.max_row or 1
    for r in range(2, max(max_row,2)+2):
        if ws.cell(r,1).value in (None,""):
            return r
    return max_row+1

def save_row(data):
    ensure_schema()
    backup_db()
    wb = load_workbook(DB_PATH)
    ws = wb[SHEET]
    vid = next_id(ws)
    data["VOC ID"] = vid
    if not data.get("Created Date"):
        data["Created Date"] = datetime.now()
    if not data.get("Updated Date"):
        data["Updated Date"] = datetime.now()
    r = first_empty_row(ws)
    for c,h in enumerate(HEADERS,1):
        ws.cell(r,c).value = data.get(h,"")
    wb.save(DB_PATH)
    wb.close()

    # verify
    wb2 = load_workbook(DB_PATH, read_only=True, data_only=False)
    ws2 = wb2[SHEET]
    saved = ws2.cell(r,1).value
    wb2.close()
    if saved != vid:
        raise RuntimeError("저장 후 검증 실패")
    return vid, r

def read_db():
    ensure_db()
    df = pd.read_excel(DB_PATH, sheet_name=SHEET)
    if "VOC ID" in df.columns:
        df = df[df["VOC ID"].notna()]
    return df


def ensure_schema():
    ensure_db()
    wb = load_workbook(DB_PATH)
    ws = wb[SHEET]
    current = [ws.cell(1,c).value for c in range(1, (ws.max_column or 1)+1)]
    changed = False
    for h in HEADERS:
        if h not in current:
            ws.cell(1, len(current)+1).value = h
            current.append(h)
            changed = True
    if changed:
        wb.save(DB_PATH)
    wb.close()

def find_row_by_voc_id(voc_id):
    ensure_schema()
    wb = load_workbook(DB_PATH, read_only=True, data_only=False)
    ws = wb[SHEET]
    found = None
    for r in range(2, (ws.max_row or 1)+1):
        if str(ws.cell(r,1).value or "").strip() == str(voc_id).strip():
            found = r
            break
    wb.close()
    return found

def read_voc(voc_id):
    ensure_schema()
    r = find_row_by_voc_id(voc_id)
    if not r:
        return None
    wb = load_workbook(DB_PATH, read_only=True, data_only=False)
    ws = wb[SHEET]
    header_map = {str(ws.cell(1,c).value): c for c in range(1,(ws.max_column or 1)+1)}
    data = {h: (ws.cell(r,header_map[h]).value if h in header_map else "") for h in HEADERS}
    wb.close()
    return data

def update_voc(voc_id, data, updated_by, update_summary):
    ensure_schema()
    backup_db()
    r = find_row_by_voc_id(voc_id)
    if not r:
        raise RuntimeError(f"{voc_id}를 찾을 수 없습니다.")
    wb = load_workbook(DB_PATH)
    ws = wb[SHEET]
    header_map = {str(ws.cell(1,c).value): c for c in range(1,(ws.max_column or 1)+1)}
    for h,v in data.items():
        c = header_map.get(h)
        if c and h not in ["VOC ID","Created By","Created Date"]:
            ws.cell(r,c).value = v
    ws.cell(r, header_map["Updated By"]).value = updated_by
    ws.cell(r, header_map["Updated Date"]).value = datetime.now()
    ws.cell(r, header_map["Last Update Summary"]).value = update_summary
    wb.save(DB_PATH)
    wb.close()
    return r

def clean_json(s):
    s = s.strip()
    s = re.sub(r"^```(?:json)?\s*","",s)
    s = re.sub(r"\s*```$","",s)
    return s.strip()

def heuristic_analysis(text):
    # API 미연결 시 최소 추출. 실제 심층 분석은 LLM 연결 권장.
    product = ""
    m = re.search(r"\b[A-Z]{2,}[A-Z0-9\-]{3,}\b", text)
    if m: product = m.group(0)

    lot = ""
    m = re.search(r"(?:批次号|批次|LOT|Lot|lot)\s*[:：]?\s*([A-Za-z0-9\-]+)", text)
    if m: lot = m.group(1)

    spec = ""
    m = re.search(r"(?:spec|SPEC|规格)\s*[:：]?\s*([^，。\n]+)", text)
    if m: spec = m.group(1).strip()

    actual = ""
    m = re.search(r"(?:实测|actual)\s*[:：]?\s*([0-9.]+)", text, re.I)
    if m: actual = m.group(1)

    qtys = re.findall(r"\d+(?:\.\d+)?\s*MM\s*[*×xX]\s*\d+(?:\.\d+)?\s*M", text, re.I)
    quantity = ", ".join(qtys)

    failure_detail = ""
    if "凹凸点" in text: failure_detail = "表面凹凸点 / 표면 요철점"
    elif "拉胶" in text: failure_detail = "拉胶 / Stringing"
    elif "起翘" in text: failure_detail = "起翘 / Lifting"

    failure_cat = "Appearance" if "凹凸点" in text or "外观" in text else ""
    impact = "客户无法出货 / 고객 출하 불가" if "无法出货" in text else ""

    requests = []
    if "退货" in text: requests.append("NG품 반품 처리")
    if "换货" in text or "合格品" in text: requests.append("합격 재고 확인 및 교환")
    if "FACA" in text.upper(): requests.append("이상 원인 분석 및 FACA 회신")

    dates = re.findall(r"\d{1,2}/\d{1,2}(?:\s+\d{1,2}[.:]\d{2})?", text)
    response_due = dates[0] if dates else ""
    faca_due = dates[1] if len(dates) > 1 else ""

    judgement = ""
    if "QTY≤3" in text.replace(" ","") and actual:
        try:
            judgement = "NG / Spec Out" if float(actual) > 3 else "OK"
        except: pass

    return {
        "product": product, "material_type": "", "lot_no": lot, "quantity": quantity,
        "process": "Customer Incoming / Material", "failure_category": failure_cat,
        "failure_detail": failure_detail, "issue_summary": re.sub(r"\s+"," ",text.strip())[:300],
        "summary_ko": "", "summary_zh": "", "summary_en": "",
        "spec": spec, "actual": actual, "judgement": judgement,
        "occurrence_condition": "", "defect_rate": "", "customer_impact": impact,
        "priority": "Critical" if "无法出货" in text or "缺货" in text else "High",
        "customer_request": "\n".join(f"{i+1}. {x}" for i,x in enumerate(requests)),
        "response_due": response_due, "faca_due": faca_due,
        "internal_action_items": "- 해당 LOT 재고 Hold\n- 동일 LOT/인접 LOT 선별\n- 대체 합격품 확인\n- 불량 Sample 확보\n- 생산/원재료 이력 Trace",
        "ai_suggested_cause": "- 원재료 표면 이상\n- 이물/Particle\n- Coating unevenness\n- Lamination/권취 압흔 가능성",
        "required_check_points": "- 동일 LOT 재고 외관 확인\n- 정상 LOT 비교\n- 폭/길이 위치별 분포 확인\n- 원재료 LOT Trace\n- 공정이력 확인\n- Sample 확대 관찰",
        "missing_information": "- 전체 불량률\n- 발생 위치 분포\n- 폭별 차이\n- 불량 Sample\n- 고객 판정 방식",
        "analysis_source":"Rule-based fallback"
    }

def build_analysis_prompt(text):
    return f"""
You are a senior NPI / Quality / Technical Sales assistant for functional adhesive tapes,
PSA, conductive fabric tape, die-cutting, lamination and electronics manufacturing.

Analyze the customer's VOC. Source may mix Simplified Chinese, Korean and English.

You must:
1. extract factual fields without inventing data;
2. distinguish FACTS from HYPOTHESES;
3. detect customer business impact and urgency;
4. split customer requests into action items and deadlines;
5. determine if the provided actual value violates the stated spec when mathematically possible;
6. propose practical internal immediate actions;
7. propose root-cause hypotheses, but NEVER label them confirmed;
8. generate required checkpoints and missing information/questions;
9. produce concise professional summaries in Korean, Simplified Chinese and English.
10. IMPORTANT UI rule: all normalized/display values must be written in KOREAN so Korean colleagues can read them.
    Keep only immutable identifiers/numbers/units/codes exactly as source (Product code, LOT, quantity, dates, spec values).
    For example:
    - material_type: "도전포 양면테이프"
    - failure_detail: "원재료 표면 요철점"
    - issue_summary: Korean
    - occurrence_condition: Korean
    - customer_impact: Korean
    - customer_request: Korean
    - internal_action_items: Korean
    - ai_suggested_cause: Korean
    - required_check_points: Korean
    - missing_information: Korean
    English/Chinese text should appear only in summary_en and summary_zh.

Return ONLY valid JSON with exactly these keys:
{{
 "product":"",
 "material_type":"",
 "lot_no":"",
 "quantity":"",
 "process":"",
 "failure_category":"",
 "failure_detail":"",
 "issue_summary":"",
 "summary_ko":"",
 "summary_zh":"",
 "summary_en":"",
 "spec":"",
 "actual":"",
 "judgement":"",
 "occurrence_condition":"",
 "defect_rate":"",
 "customer_impact":"",
 "priority":"",
 "customer_request":[""],
 "response_due":"",
 "faca_due":"",
 "internal_action_items":[""],
 "ai_suggested_cause":[""],
 "required_check_points":[""],
 "missing_information":[""]
}}

Allowed process labels:
{PROCESSES}

Allowed failure_category labels:
{FAILURES}

Allowed priority: Low, Medium, High, Critical.

Important:
- "Critical" is appropriate for line stop, shipment blocked, customer shortage, or similarly urgent business impact.
- Preserve exact product codes, LOT, quantities, spec values and deadlines.
- Customer Request should be split into independent requests.
- If date year is absent, keep it as written instead of guessing a year.
- Chinese summaries must use Simplified Chinese.
- Use Korean technical terminology suitable for an NPI sales engineer.

VOC SOURCE:
{text}
"""

def call_openai(prompt, api_key, model):
    if OpenAI is None:
        raise RuntimeError("openai 패키지가 설치되지 않았습니다.")
    client = OpenAI(api_key=api_key)
    resp = client.responses.create(model=model, input=prompt)
    return resp.output_text

def call_gemini(prompt, api_key, model):
    if genai is None:
        raise RuntimeError("google-genai 패키지가 설치되지 않았습니다.")
    client = genai.Client(api_key=api_key)
    # generate_content는 stable SDK path로 사용
    resp = client.models.generate_content(model=model, contents=prompt)
    return resp.text

def call_openrouter(prompt, api_key, model):
    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.1
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost:8501",
            "X-Title": "VOC Intelligence Assistant"
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as res:
            data = json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"OpenRouter API 오류 {e.code}: {body[:500]}")
    return data["choices"][0]["message"]["content"]

def call_provider(provider, prompt, api_key, model):
    if provider == "OpenAI":
        return call_openai(prompt, api_key, model)
    if provider == "Google Gemini":
        return call_gemini(prompt, api_key, model)
    if provider == "OpenRouter Free":
        return call_openrouter(prompt, api_key, model)
    raise RuntimeError("지원하지 않는 AI Provider입니다.")

def is_transient_api_error(exc):
    """일시적 서버 혼잡/속도제한/타임아웃 계열인지 판별."""
    msg = str(exc).lower()
    transient_tokens = [
        "503", "502", "504", "500", "429",
        "unavailable", "high demand", "overloaded",
        "rate limit", "resource_exhausted",
        "timeout", "timed out", "temporarily"
    ]
    return any(token in msg for token in transient_tokens)

def alternate_model(provider, model):
    """같은 Provider 내에서 가능한 경우 대체 모델 반환."""
    if provider == "Google Gemini":
        candidates = ["gemini-flash-latest", "gemini-3.6-flash"]
        for candidate in candidates:
            if candidate != model:
                return candidate
    return None

def deep_analyze(text, provider, api_key, model):
    if provider == "Local / No API" or not api_key:
        result = heuristic_analysis(text)
        result["analysis_source"] = "Local / Rule-based"
        return result

    prompt = build_analysis_prompt(text)
    attempts = []
    last_error = None

    # 1) 선택한 모델 최대 2회 재시도
    for attempt in range(2):
        try:
            raw = call_provider(provider, prompt, api_key, model)
            result = json.loads(clean_json(raw))
            result["analysis_source"] = f"{provider} ({model})"
            if attempt > 0:
                result["analysis_warning"] = "AI 서버 혼잡으로 재시도 후 분석에 성공했습니다."
            return result
        except Exception as e:
            last_error = e
            attempts.append(f"{model}: {str(e)[:180]}")
            if not is_transient_api_error(e):
                raise
            if attempt == 0:
                time.sleep(2)

    # 2) 같은 Provider의 다른 모델로 1회 자동 전환
    alt = alternate_model(provider, model)
    if alt:
        try:
            raw = call_provider(provider, prompt, api_key, alt)
            result = json.loads(clean_json(raw))
            result["analysis_source"] = f"{provider} ({alt}) · Auto Fallback"
            result["analysis_warning"] = (
                f"선택 모델({model})이 일시적으로 혼잡하여 "
                f"대체 모델({alt})로 자동 분석했습니다."
            )
            return result
        except Exception as e:
            last_error = e
            attempts.append(f"{alt}: {str(e)[:180]}")
            if not is_transient_api_error(e):
                raise

    # 3) 일시적 장애가 계속되면 Local 분석으로 안전하게 폴백
    if last_error is not None and is_transient_api_error(last_error):
        result = heuristic_analysis(text)
        result["analysis_source"] = "Local / Rule-based · Emergency Fallback"
        result["analysis_warning"] = (
            "AI 서버가 일시적으로 혼잡하여 규칙 기반 분석 결과를 표시합니다. "
            "중요 VOC는 잠시 후 AI로 다시 분석해 주세요."
        )
        result["api_attempt_log"] = attempts
        return result

    raise last_error

def translate_with_provider(text, provider, api_key, model):
    if provider == "Local / No API" or not api_key:
        raise RuntimeError("자연어 한·중·영 변환은 AI Provider와 API Key가 필요합니다.")
    prompt = f"""Translate this NPI/VOC technical text into Korean, Simplified Chinese, and concise professional English.
Preserve product codes, LOT, numbers, dates, units and technical meaning.
Return ONLY valid JSON: {{"ko":"","zh":"","en":""}}
TEXT:
{text}"""
    raw = call_provider(provider, prompt, api_key, model)
    return json.loads(clean_json(raw))

def list_to_text(v, numbered=False):
    if isinstance(v,list):
        if numbered:
            return "\n".join(f"{i+1}. {x}" for i,x in enumerate(v) if x)
        return "\n".join(f"- {x}" for x in v if x)
    return v or ""

ensure_db()



PROCESS_DISPLAY = {
    "": "",
    "Customer Incoming / Material": "고객 입고/원자재",
    "Coating": "코팅",
    "Drying/Oven": "건조/오븐",
    "Lamination": "합지/Lamination",
    "Slitting": "슬리팅",
    "Die-cut": "타발/Die-cut",
    "Waste Stripping": "배폐/Waste Stripping",
    "Peeling/Release": "이형/박리",
    "Assembly": "조립",
    "Reliability Test": "신뢰성 시험",
    "Customer Process": "고객 공정",
    "Other": "기타",
}
FAILURE_DISPLAY = {
    "": "",
    "Appearance": "외관",
    "拉胶 / Stringing": "라교/Stringing",
    "溢胶 / Adhesive Overflow": "접착제 넘침/Overflow",
    "残胶 / Residue": "잔사/Residue",
    "起翘 / Lifting": "들뜸/Lifting",
    "Curl": "컬/Curl",
    "Bubble": "기포/Bubble",
    "Wrinkle": "주름/Wrinkle",
    "Scratch": "스크래치/Scratch",
    "Foreign Material": "이물/Foreign Material",
    "Release Force Low": "이형력 낮음",
    "Release Force High": "이형력 높음",
    "Peel Low": "박리력 낮음",
    "High-temp Peel": "고온 박리",
    "Dimension": "치수",
    "Other": "기타",
}
PRIORITY_DISPLAY = {"Low":"낮음","Medium":"보통","High":"높음","Critical":"긴급"}

UI_LABELS = {
    "접수일": "접수일 / Received Date / 接收日期",
    "고객사": "고객사 / Customer / 客户",
    "담당자": "담당자 / DRI / 负责人",
    "제품명": "제품명 / Product / 产品",
    "제품유형": "제품 유형 / Material · Product Type / 材料·产品类型",
    "제품구조": "제품 구조 / Product Structure / 产品结构",
    "LOT": "LOT No. / Lot No. / 批次号",
    "수량": "수량 / Quantity / 数量",
    "공정": "공정 / Process / 工序",
    "불량분류": "불량 분류 / Failure Category / 不良分类",
    "불량상세": "불량 상세 / Failure Detail / 不良详情",
    "규격": "규격 / SPEC / 规格",
    "실측": "실측 / Actual / 实测",
    "판정": "판정 / Judgement / 判定",
    "우선순위": "우선순위 / Priority / 优先级",
    "이슈요약": "이슈 요약 / Issue Summary / 问题摘要",
    "발생조건": "발생 조건 / Occurrence Condition / 发生条件",
    "불량률": "불량률 / Defect Rate / 不良率",
    "고객영향": "고객 영향 / Customer Impact / 客户影响",
    "고객요청": "고객 요청 / Customer Request / 客户要求",
    "회신기한": "회신 기한 / Response Due / 回复期限",
    "FACA기한": "FACA 기한 / FACA Due / FACA期限",
    "내부조치": "즉시 내부 조치 / Internal Immediate Action / 内部即时措施",
    "체크포인트": "확인 필요 항목 / Required Check Points / 必要确认项目",
    "원인가설": "AI 원인 가설 / AI Suggested Cause / AI原因假设",
    "추가정보": "추가 확인 정보 / Missing Information / 待确认信息",
    "한국어요약": "한국어 요약 / Korean Summary / 韩文摘要",
    "중국어요약": "중국어 요약 / Chinese Summary / 中文摘要",
    "영어요약": "영어 요약 / English Summary / 英文摘要",
    "확정원인": "확정 원인 / Confirmed Root Cause / 确认根因",
    "유출원인": "유출 원인 / Escape Cause / 流出原因",
    "개선조치": "개선 조치 / Corrective Action / 改善措施",
    "검증결과": "검증 결과 / Verification Result / 验证结果",
    "상태": "상태 / Status / 状态",
    "첨부": "사진·링크 / Picture · Link / 图片·链接",
    "비고": "비고 / Remark / 备注",
}

APP_SETTINGS = load_settings()

st.set_page_config(
    page_title="VOC Intelligence",
    page_icon=str(BASE / "assets" / "voc_roman_icon.png"),
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .block-container {padding-top: 1.35rem; padding-bottom: 2rem; max-width: 1500px;}
    [data-testid="stSidebar"] {border-right: 1px solid rgba(49, 51, 63, 0.12);}
    .voc-hero {
        padding: 1.35rem 1.55rem;
        border: 1px solid rgba(49, 51, 63, 0.12);
        border-radius: 18px;
        margin-bottom: 1.15rem;
        background: rgba(250, 250, 252, 0.72);
    }
    .voc-title {
        font-size: 2.0rem;
        font-weight: 760;
        margin: 0;
        line-height: 1.15;
        letter-spacing: -0.02em;
    }
    .voc-subtitle {
        margin-top: .35rem;
        font-size: 1rem;
        color: rgba(49, 51, 63, 0.72);
    }
    .voc-badge {
        display: inline-block;
        padding: .22rem .62rem;
        border-radius: 999px;
        font-size: .78rem;
        border: 1px solid rgba(49, 51, 63, 0.15);
        margin-top: .6rem;
    }
    .voc-hero {box-shadow: 0 2px 12px rgba(16,24,40,.04);}
    div[data-testid="stMetric"] {border:1px solid rgba(49,51,63,.10); padding:12px 14px; border-radius:14px; background:rgba(250,250,252,.65);}
    div[data-testid="stDataFrame"] {border:1px solid rgba(49,51,63,.10); border-radius:12px; overflow:hidden;}
</style>
""", unsafe_allow_html=True)

hero_left, hero_right = st.columns([7, 1])
with hero_left:
    st.markdown("""
    <div class="voc-hero">
        <div class="voc-title">VOC Intelligence</div>
        <div class="voc-subtitle">
            AI-assisted VOC Management · Analysis · FACA Support
        </div>
        <div class="voc-badge">NPI Sales Workflow · V3.5.6 Business Edition</div>
    </div>
    """, unsafe_allow_html=True)
with hero_right:
    st.image(str(BASE / "assets" / "voc_roman_icon.png"), width=86)

# session defaults
for k in ["translation_message","analysis_source"]:
    if k not in st.session_state: st.session_state[k] = ""

with st.sidebar:
    st.image(str(BASE / "assets" / "voc_roman_icon.png"), width=72)
    st.markdown("### VOC Intelligence")
    st.caption("VOC Management & AI Analysis")
    st.divider()

    st.subheader("사용자 · 설정")
    current_user = st.text_input(
        "현재 사용자 / Current User / 当前用户",
        value=st.session_state.get("current_user", APP_SETTINGS.get("current_user","Jongin"))
    )
    st.session_state["current_user"] = current_user.strip() or "Unknown"
    st.caption("신규 등록/수정 시 Created By, Updated By에 기록됩니다.")
    st.divider()

    st.subheader("AI 연결")
    provider_options = ["Google Gemini", "OpenRouter Free", "OpenAI", "Local / No API"]
    saved_provider = APP_SETTINGS.get("provider","Google Gemini")
    provider = st.selectbox(
        "AI Provider", provider_options,
        index=provider_options.index(saved_provider) if saved_provider in provider_options else 0,
        help="무료 테스트는 Gemini 또는 OpenRouter Free를 추천합니다."
    )

    if provider == "Google Gemini":
        saved_key = load_local_key(provider) or os.getenv("GEMINI_API_KEY","")
        api_key = st.text_input("Gemini API Key", value=saved_key, type="password")
        models = ["gemini-flash-latest", "gemini-3.6-flash"]
        saved_model = APP_SETTINGS.get("gemini_model",models[0])
        model = st.selectbox("Model", models, index=models.index(saved_model) if saved_model in models else 0)
        st.caption("무료 Tier 사용 가능 여부/한도는 Google AI Studio 계정 상태에 따라 달라질 수 있습니다.")
    elif provider == "OpenRouter Free":
        saved_key = load_local_key(provider) or os.getenv("OPENROUTER_API_KEY","")
        api_key = st.text_input("OpenRouter API Key", value=saved_key, type="password")
        models = ["openrouter/free"]
        model = st.selectbox("Model", models, index=0)
        st.caption("openrouter/free는 사용 가능한 무료 모델 중 하나로 자동 라우팅됩니다.")
    elif provider == "OpenAI":
        saved_key = load_local_key(provider) or os.getenv("OPENAI_API_KEY","")
        api_key = st.text_input("OpenAI API Key", value=saved_key, type="password")
        models = ["gpt-5.6-luna","gpt-5.6-terra","gpt-5.6-sol"]
        saved_model = APP_SETTINGS.get("openai_model",models[0])
        model = st.selectbox("Model", models, index=models.index(saved_model) if saved_model in models else 0)
        st.caption("Luna: 저비용 / Terra: 균형 / Sol: 복잡한 분석")
    else:
        api_key = ""
        model = ""
        st.info("API 없이 규칙 기반 분석만 사용합니다.")

    st.divider()
    st.subheader("설정 저장")
    remember_key = st.checkbox("이 Mac에 현재 API Key 저장", value=bool(load_local_key(provider)) if provider != "Local / No API" else False, disabled=(provider == "Local / No API"))
    if st.button("💾 현재 설정 저장", width='stretch'):
        cfg = {
            "current_user": st.session_state.get("current_user","Unknown"),
            "provider": provider,
            "gemini_model": model if provider == "Google Gemini" else APP_SETTINGS.get("gemini_model","gemini-flash-latest"),
            "openrouter_model": model if provider == "OpenRouter Free" else APP_SETTINGS.get("openrouter_model","openrouter/free"),
            "openai_model": model if provider == "OpenAI" else APP_SETTINGS.get("openai_model","gpt-5.6-luna"),
        }
        save_settings(cfg)
        if provider != "Local / No API":
            save_local_key(provider, api_key if remember_key else "")
        st.success("설정을 저장했습니다. 다음 실행부터 자동 적용됩니다.")
    st.caption("API Key는 이 프로젝트 폴더의 로컬 숨김 파일에만 저장됩니다. 외부 공유 ZIP에는 Key를 넣지 마세요.")

    st.divider()
    st.subheader("DB")
    st.metric("저장 VOC", count_records())
    st.code(str(DB_PATH))
    with open(DB_PATH,"rb") as f:
        st.download_button("📥 실제 Excel DB 다운로드", f.read(), "VOC_Master_AI.xlsx")

tabs = st.tabs(["📊 Dashboard","🧠 VOC 심층 분석","✏️ 기존 VOC 수정","📚 최근 VOC","🌐 한·중·영","ℹ️ 가이드"])


with tabs[0]:
    st.markdown("## Dashboard")
    st.markdown("""
    <div style="margin:-0.25rem 0 0.8rem 0;padding:0.55rem 0.8rem;border-radius:10px;
    background:rgba(46,144,250,.06);border:1px solid rgba(46,144,250,.12);
    font-size:.82rem;color:#475467;">
    Business Edition · Today's Action + Quick Status Update
    </div>
    """, unsafe_allow_html=True)
    st.caption("오늘 처리할 VOC와 전체 진행 현황을 한눈에 확인합니다. · V3.5.6")
    df_dash = read_db()

    if len(df_dash) == 0:
        st.info("아직 등록된 VOC가 없습니다.")
    else:
        total = len(df_dash)
        if "Status" in df_dash.columns:
            status_series = df_dash["Status"].fillna("").astype(str)
            status_norm = status_series.str.strip().str.lower()

            closed_count = int(status_norm.eq("closed").sum())
            pending_count = int(status_norm.isin(["pending","pengding","대기","보류"]).sum())
            drop_count = int(status_norm.isin(["drop","dropped","취소","중단"]).sum())
            overdue_status_count = int(status_norm.isin(["over due","overdue","기한초과"]).sum())

            # Open은 Closed / Drop을 제외한 현재 진행성 VOC 전체
            open_count = int((~status_norm.isin(["closed","drop","dropped","취소","중단"])).sum())
        else:
            closed_count = 0
            pending_count = 0
            drop_count = 0
            overdue_status_count = 0
            open_count = total

        due_rows = []
        due_count = 0
        faca_overdue_count = 0
        if "FACA Due" in df_dash.columns:
            for _, row in df_dash.iterrows():
                row_status = str(row.get("Status","")).strip().lower()
                if row_status in ["closed","drop","dropped","취소","중단"]:
                    continue
                state, dd = due_state(row.get("FACA Due",""))
                if state == "overdue":
                    faca_overdue_count += 1
                if state in ["soon","overdue"]:
                    due_count += 1
                    due_rows.append({
                        "VOC ID": row.get("VOC ID",""),
                        "Customer": row.get("Customer",""),
                        "Product": row.get("Product",""),
                        "Priority": row.get("Priority",""),
                        "FACA Due": row.get("FACA Due",""),
                        "Status": row.get("Status",""),
                        "Due State": "기한초과" if state == "overdue" else "3일 이내"
                    })

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("전체 VOC", total)
        c2.metric("Open", open_count)
        c3.metric("Pending", pending_count)
        c4.metric("Over Due", max(overdue_status_count, faca_overdue_count))
        c5.metric("Drop", drop_count)
        c6.metric("Closed", closed_count)

        st.caption("Over Due는 Status가 Over Due이거나, 미종결 VOC의 FACA Due가 지난 건을 반영합니다.")


        # V3.5 Today's Action — Response/FACA due and overdue items
        st.divider()
        st.markdown("#### 🔔 Today's Action")
        action_rows = []
        today_ts = pd.Timestamp(date.today())
        for _, row in df_dash.iterrows():
            rs = str(row.get("Status","")).strip().lower()
            if rs in ["closed","drop","dropped","취소","중단"]:
                continue
            alerts = []
            urgency = 99
            for field, label in [("Response Due","Response"),("FACA Due","FACA")]:
                raw = row.get(field,"")
                dt = pd.to_datetime(raw, errors="coerce")
                if pd.notna(dt):
                    days = (dt.normalize() - today_ts).days
                    if days < 0:
                        alerts.append(f"{label} Due D+{abs(days)}")
                        urgency = min(urgency, -100 + days)
                    elif days == 0:
                        alerts.append(f"{label} Due D-0")
                        urgency = min(urgency, 0)
                    elif days <= 3:
                        alerts.append(f"{label} Due D-{days}")
                        urgency = min(urgency, days)
            if rs in ["over due","overdue","기한초과"] and not alerts:
                alerts.append("Status: Over Due")
                urgency = -90
            if alerts:
                action_rows.append({
                    "VOC ID": row.get("VOC ID",""), "Customer": row.get("Customer",""),
                    "Product": row.get("Product",""), "Action": " · ".join(alerts),
                    "Status": row.get("Status",""), "_urgency": urgency
                })
        if action_rows:
            adf = pd.DataFrame(action_rows).sort_values("_urgency").drop(columns=["_urgency"]).head(12)
            st.dataframe(adf, width='stretch', hide_index=True)
        else:
            st.success("오늘 기준 기한 초과 또는 3일 이내 처리 예정 VOC가 없습니다.")

        # Quick Status Update — status only, with audit trail
        with st.expander("⚡ Quick Status Update · Dashboard에서 상태 빠르게 변경"):
            active = df_dash.copy()
            if "VOC ID" in active.columns:
                ids = [str(x) for x in active["VOC ID"].dropna().tolist() if str(x).strip()]
                if ids:
                    q1,q2 = st.columns([2,1])
                    with q1:
                        quick_id = st.selectbox("VOC ID", ids, key="v35_quick_id")
                    current_row = active[active["VOC ID"].astype(str)==str(quick_id)].iloc[-1]
                    cur_status = str(current_row.get("Status","Open") or "Open")
                    with q2:
                        quick_status = st.selectbox("Status", STATUSES, index=STATUSES.index(cur_status) if cur_status in STATUSES else 0, key="v35_quick_status")
                    st.caption(f"{current_row.get('Customer','')} · {current_row.get('Product','')} · 현재 상태: {cur_status}")
                    if st.button("상태 업데이트", key="v35_quick_save", type="primary"):
                        try:
                            close_date = date.today() if quick_status == "Closed" else current_row.get("Close Date","")
                            update_voc(quick_id, {"Status":quick_status,"Close Date":close_date}, st.session_state.get("current_user","Unknown"), f"Dashboard Quick Status: {cur_status} → {quick_status}")
                            st.success(f"{quick_id}: {cur_status} → {quick_status} 업데이트 완료")
                            st.rerun()
                        except Exception as e:
                            st.error(f"상태 업데이트 실패: {e}")

        st.divider()
        l1,r1 = st.columns(2)
        with l1:
            st.markdown("#### 고객별 VOC")
            if "Customer" in df_dash.columns:
                s = df_dash["Customer"].fillna("미지정").replace("","미지정").value_counts().head(10)
                st.bar_chart(s)
        with r1:
            st.markdown("#### 불량 유형별 VOC")
            if "Failure Category" in df_dash.columns:
                s = df_dash["Failure Category"].fillna("미지정").replace("","미지정").value_counts().head(10)
                st.bar_chart(s)

        st.divider()
        l2,r2 = st.columns(2)
        with l2:
            st.markdown("#### 제품별 VOC")
            if "Product" in df_dash.columns:
                pc = df_dash["Product"].fillna("미지정").replace("","미지정").value_counts().head(10).rename_axis("Product").reset_index(name="Count")
                st.dataframe(pc, width='stretch', hide_index=True)
        with r2:
            st.markdown("#### 상태별 현황")
            if "Status" in df_dash.columns:
                sc = df_dash["Status"].fillna("미지정").replace("","미지정").value_counts().rename_axis("Status").reset_index(name="Count")
                st.dataframe(sc, width='stretch', hide_index=True)

        st.divider()
        st.markdown("#### FACA Due 관리")
        if due_rows:
            due_df = pd.DataFrame(due_rows)
            st.dataframe(due_df, width='stretch', hide_index=True)
        else:
            st.success("현재 FACA 기한 초과 또는 3일 이내 예정 VOC가 없습니다.")

        st.divider()
        st.markdown("#### Open Issue")
        open_df = df_dash.copy()
        if "Status" in open_df.columns:
            open_df = open_df[~open_df["Status"].fillna("").astype(str).str.strip().str.lower().isin(["closed","drop","dropped","취소","중단"])]
        cols = [c for c in ["VOC ID","Customer","Product","Failure Category","Priority","Response Due","FACA Due","Status","Updated Date"] if c in open_df.columns]
        if len(open_df):
            st.dataframe(open_df[cols].tail(50), width='stretch', hide_index=True)
        else:
            st.success("Open VOC가 없습니다.")

        st.divider()
        st.markdown("#### 최근 수정 VOC")
        if "Updated Date" in df_dash.columns:
            recent = df_dash.copy()
            recent["_sort"] = pd.to_datetime(recent["Updated Date"], errors="coerce")
            recent = recent.sort_values("_sort", ascending=False).head(10)
            rcols = [c for c in ["VOC ID","Customer","Product","Status","Updated By","Updated Date","Last Update Summary"] if c in recent.columns]
            st.dataframe(recent[rcols], width='stretch', hide_index=True)


with tabs[1]:
    st.subheader("1. 고객 VOC 원문")
    st.text_area("VOC Original Text", key="voc_original", height=220,
                 placeholder="고객 메일/WeChat/VOC 내용을 그대로 붙여넣으세요.")

    def run_analysis():
        text = st.session_state.get("voc_original","").strip()
        if not text:
            st.session_state["analysis_error"] = "VOC 원문을 입력하세요."
            return
        try:
            r = deep_analyze(text, provider, api_key, model)
            mapping = {
                "product":"product","material_type":"material_type","lot_no":"lot_no","quantity":"quantity",
                "process":"process","failure_category":"failure_category","failure_detail":"failure_detail",
                "issue_summary":"issue_summary","summary_ko":"summary_ko","summary_zh":"summary_zh","summary_en":"summary_en",
                "spec":"spec","actual":"actual","judgement":"judgement","occurrence_condition":"occurrence_condition",
                "defect_rate":"defect_rate","customer_impact":"customer_impact","priority":"priority",
                "response_due":"response_due","faca_due":"faca_due"
            }
            for src,dst in mapping.items():
                val = r.get(src,"")
                if dst == "process" and val not in PROCESSES: val = ""
                if dst == "failure_category" and val not in FAILURES: val = ""
                if dst == "priority" and val not in PRIORITIES: val = "High"
                st.session_state[dst] = val

            st.session_state["customer_request"] = list_to_text(r.get("customer_request",[]), numbered=True)
            st.session_state["internal_action_items"] = list_to_text(r.get("internal_action_items",[]))
            st.session_state["ai_suggested_cause"] = list_to_text(r.get("ai_suggested_cause",[]))
            st.session_state["required_check_points"] = list_to_text(r.get("required_check_points",[]))
            st.session_state["missing_information"] = list_to_text(r.get("missing_information",[]))
            st.session_state["analysis_source"] = r.get("analysis_source","")
            st.session_state["analysis_warning"] = r.get("analysis_warning","")
            st.session_state["analysis_error"] = ""
        except Exception as e:
            st.session_state["analysis_warning"] = ""
            st.session_state["analysis_error"] = str(e)

    st.button("🧠 VOC 심층 분석", type="primary", on_click=run_analysis, width='stretch')

    if st.session_state.get("analysis_error"):
        err = st.session_state["analysis_error"]
        if is_transient_api_error(RuntimeError(err)):
            st.warning("⚠️ AI 서버가 일시적으로 혼잡합니다. 잠시 후 다시 시도해 주세요.")
        else:
            st.error(f"AI 분석 오류: {err}")
    elif st.session_state.get("analysis_source"):
        st.success(f"분석 완료 · {st.session_state['analysis_source']}")
        if st.session_state.get("analysis_warning"):
            st.warning(st.session_state["analysis_warning"])

    # Headline intelligence cards
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("우선순위 / Priority / 优先级", st.session_state.get("priority","-") or "-")
    c2.metric("판정 / Judgement / 判定", st.session_state.get("judgement","-") or "-")
    c3.metric("회신 기한 / Response Due / 回复期限", st.session_state.get("response_due","-") or "-")
    c4.metric("FACA 기한 / FACA Due / FACA期限", st.session_state.get("faca_due","-") or "-")

    st.subheader("2. 핵심 추출 결과")
    a,b,c = st.columns(3)
    with a:
        st.date_input(UI_LABELS["접수일"], date.today(), key="voc_date")
        st.text_input(UI_LABELS["고객사"], key="customer")
        st.text_input(UI_LABELS["담당자"], key="dri")
        st.text_input(UI_LABELS["제품명"], key="product")
        st.text_input(UI_LABELS["제품유형"], key="material_type")
    with b:
        st.text_input(UI_LABELS["제품구조"], key="product_structure")
        st.text_input(UI_LABELS["LOT"], key="lot_no")
        st.text_input(UI_LABELS["수량"], key="quantity")
        st.selectbox(UI_LABELS["공정"], PROCESSES, key="process", format_func=lambda x: PROCESS_DISPLAY.get(x, x))
        st.selectbox(UI_LABELS["불량분류"], FAILURES, key="failure_category", format_func=lambda x: FAILURE_DISPLAY.get(x, x))
    with c:
        st.text_input(UI_LABELS["불량상세"], key="failure_detail")
        st.text_input(UI_LABELS["규격"], key="spec")
        st.text_input(UI_LABELS["실측"], key="actual")
        st.text_input(UI_LABELS["판정"], key="judgement")
        st.selectbox(UI_LABELS["우선순위"], PRIORITIES, key="priority", format_func=lambda x: PRIORITY_DISPLAY.get(x, x))

    st.text_area(UI_LABELS["이슈요약"], key="issue_summary", height=90)
    st.text_input(UI_LABELS["발생조건"], key="occurrence_condition")
    st.text_input(UI_LABELS["불량률"], key="defect_rate")
    st.text_area(UI_LABELS["고객영향"], key="customer_impact", height=80)

    st.subheader("3. 고객 요구 & Deadline")
    x,y,z = st.columns([2,1,1])
    with x: st.text_area(UI_LABELS["고객요청"], key="customer_request", height=150)
    with y: st.text_input(UI_LABELS["회신기한"], key="response_due")
    with z: st.text_input(UI_LABELS["FACA기한"], key="faca_due")

    st.subheader("4. NPI 대응 Assistant")
    p,q = st.columns(2)
    with p:
        st.text_area("🚨 " + UI_LABELS["내부조치"], key="internal_action_items", height=190)
        st.text_area("🔍 " + UI_LABELS["체크포인트"], key="required_check_points", height=190)
    with q:
        st.text_area("💡 " + UI_LABELS["원인가설"], key="ai_suggested_cause", height=190)
        st.text_area("❓ " + UI_LABELS["추가정보"], key="missing_information", height=190)

    st.subheader("5. 한 · 중 · 영 요약")
    l1,l2,l3 = st.columns(3)
    with l1: st.text_area(UI_LABELS["한국어요약"], key="summary_ko", height=150)
    with l2: st.text_area(UI_LABELS["중국어요약"], key="summary_zh", height=150)
    with l3: st.text_area(UI_LABELS["영어요약"], key="summary_en", height=150)

    with st.expander("6. FACA / 확정 분석 — 사람 검토 후 입력"):
        st.text_area(UI_LABELS["확정원인"], key="confirmed_root")
        st.text_area(UI_LABELS["유출원인"], key="escape_cause")
        st.text_area(UI_LABELS["개선조치"], key="corrective_action")
        st.text_area(UI_LABELS["검증결과"], key="verification_result")
        st.text_input("FACA No.", key="faca_no")
        st.selectbox(UI_LABELS["상태"], STATUSES, key="status")
        st.text_input(UI_LABELS["첨부"], key="picture_link")
        st.text_area(UI_LABELS["비고"], key="remark")

    if st.button("💾 검토 완료 → Excel DB 저장", width='stretch'):
        row = {
            "Date": st.session_state["voc_date"],
            "Customer": st.session_state.get("customer",""),
            "DRI": st.session_state.get("dri",""),
            "Product": st.session_state.get("product",""),
            "Material / Product Type": st.session_state.get("material_type",""),
            "Product Structure": st.session_state.get("product_structure",""),
            "Lot No.": st.session_state.get("lot_no",""),
            "Quantity": st.session_state.get("quantity",""),
            "Process": st.session_state.get("process",""),
            "Failure Category": st.session_state.get("failure_category",""),
            "Failure Detail": st.session_state.get("failure_detail",""),
            "VOC Original Text": st.session_state.get("voc_original",""),
            "Issue Summary": st.session_state.get("issue_summary",""),
            "Summary_KO": st.session_state.get("summary_ko",""),
            "Summary_ZH": st.session_state.get("summary_zh",""),
            "Summary_EN": st.session_state.get("summary_en",""),
            "SPEC": st.session_state.get("spec",""),
            "Actual": st.session_state.get("actual",""),
            "Judgement": st.session_state.get("judgement",""),
            "Occurrence Condition": st.session_state.get("occurrence_condition",""),
            "Defect Rate": st.session_state.get("defect_rate",""),
            "Customer Impact": st.session_state.get("customer_impact",""),
            "Priority": st.session_state.get("priority","High"),
            "Customer Request": st.session_state.get("customer_request",""),
            "Response Due": st.session_state.get("response_due",""),
            "FACA Due": st.session_state.get("faca_due",""),
            "Internal Action Items": st.session_state.get("internal_action_items",""),
            "AI Suggested Cause": st.session_state.get("ai_suggested_cause",""),
            "Required Check Points": st.session_state.get("required_check_points",""),
            "Missing Information": st.session_state.get("missing_information",""),
            "Confirmed Root Cause": st.session_state.get("confirmed_root",""),
            "Escape Cause": st.session_state.get("escape_cause",""),
            "Corrective Action": st.session_state.get("corrective_action",""),
            "Verification Result": st.session_state.get("verification_result",""),
            "FACA No.": st.session_state.get("faca_no",""),
            "Status": st.session_state.get("status","Open"),
            "Close Date": date.today() if st.session_state.get("status") == "Closed" else "",
            "Picture / Link": st.session_state.get("picture_link",""),
            "Remark": st.session_state.get("remark",""),
            "Created By": st.session_state.get("current_user","Unknown"),
            "Created Date": datetime.now(),
            "Updated By": st.session_state.get("current_user","Unknown"),
            "Updated Date": datetime.now(),
            "Last Update Summary": "신규 VOC 등록"
        }
        try:
            vid, r = save_row(row)
            st.success(f"✅ {vid} 저장 및 재검증 완료 · Excel {r}행")
        except Exception as e:
            st.error(f"저장 실패: {e}")

with tabs[2]:
    st.subheader("기존 VOC 검색 · 수정")
    df_edit = read_db()
    if len(df_edit) == 0:
        st.info("저장된 VOC가 없습니다.")
    else:
        search = st.text_input("VOC 검색 / Search / 搜索", placeholder="VOC ID / Product / LOT / Customer")
        result = df_edit.copy()
        if search:
            mask = result.astype(str).apply(lambda row: row.str.contains(search, case=False, na=False).any(), axis=1)
            result = result[mask]
        voc_options = result["VOC ID"].astype(str).tolist() if "VOC ID" in result.columns else []
        if not voc_options:
            st.warning("검색 결과가 없습니다.")
        else:
            selected_voc = st.selectbox("수정할 VOC 선택 / Select VOC / 选择VOC", voc_options)
            rec = read_voc(selected_voc)
            if rec:
                c1,c2,c3,c4 = st.columns(4)
                c1.metric("VOC ID", selected_voc)
                c2.metric("상태 / Status / 状态", rec.get("Status","") or "-")
                c3.metric("작성자 / Created By / 创建者", rec.get("Created By","") or "-")
                c4.metric("최종 수정자 / Updated By / 更新者", rec.get("Updated By","") or "-")
                st.caption(f"Created: {rec.get('Created Date','')}  |  Updated: {rec.get('Updated Date','')}")
                e1,e2,e3 = st.columns(3)
                with e1:
                    ed_customer = st.text_input(UI_LABELS["고객사"], value=str(rec.get("Customer","") or ""), key="ed_customer")
                    ed_dri = st.text_input(UI_LABELS["담당자"], value=str(rec.get("DRI","") or ""), key="ed_dri")
                    ed_product = st.text_input(UI_LABELS["제품명"], value=str(rec.get("Product","") or ""), key="ed_product")
                    ed_lot = st.text_input(UI_LABELS["LOT"], value=str(rec.get("Lot No.","") or ""), key="ed_lot")
                with e2:
                    current_process = rec.get("Process","") or ""
                    pidx = PROCESSES.index(current_process) if current_process in PROCESSES else 0
                    ed_process = st.selectbox(UI_LABELS["공정"], PROCESSES, index=pidx, format_func=lambda x: PROCESS_DISPLAY.get(x,x), key="ed_process")
                    current_failure = rec.get("Failure Category","") or ""
                    fidx = FAILURES.index(current_failure) if current_failure in FAILURES else 0
                    ed_failure = st.selectbox(UI_LABELS["불량분류"], FAILURES, index=fidx, format_func=lambda x: FAILURE_DISPLAY.get(x,x), key="ed_failure")
                    current_priority = rec.get("Priority","High") or "High"
                    pridx = PRIORITIES.index(current_priority) if current_priority in PRIORITIES else 2
                    ed_priority = st.selectbox(UI_LABELS["우선순위"], PRIORITIES, index=pridx, format_func=lambda x: PRIORITY_DISPLAY.get(x,x), key="ed_priority")
                    status_value = rec.get("Status","Open") or "Open"
                    sidx = STATUSES.index(status_value) if status_value in STATUSES else 0
                    ed_status = st.selectbox(UI_LABELS["상태"], STATUSES, index=sidx, key="ed_status")
                with e3:
                    ed_response_due = st.text_input(UI_LABELS["회신기한"], value=str(rec.get("Response Due","") or ""), key="ed_response_due")
                    ed_faca_due = st.text_input(UI_LABELS["FACA기한"], value=str(rec.get("FACA Due","") or ""), key="ed_faca_due")
                    ed_faca_no = st.text_input("FACA No.", value=str(rec.get("FACA No.","") or ""), key="ed_faca_no")
                ed_summary = st.text_area(UI_LABELS["이슈요약"], value=str(rec.get("Issue Summary","") or ""), height=100, key="ed_summary")
                ed_request = st.text_area(UI_LABELS["고객요청"], value=str(rec.get("Customer Request","") or ""), height=130, key="ed_request")
                ed_actions = st.text_area(UI_LABELS["내부조치"], value=str(rec.get("Internal Action Items","") or ""), height=150, key="ed_actions")
                ed_checks = st.text_area(UI_LABELS["체크포인트"], value=str(rec.get("Required Check Points","") or ""), height=150, key="ed_checks")
                ed_root = st.text_area(UI_LABELS["확정원인"], value=str(rec.get("Confirmed Root Cause","") or ""), height=120, key="ed_root")
                ed_escape = st.text_area(UI_LABELS["유출원인"], value=str(rec.get("Escape Cause","") or ""), height=100, key="ed_escape")
                ed_corrective = st.text_area(UI_LABELS["개선조치"], value=str(rec.get("Corrective Action","") or ""), height=130, key="ed_corrective")
                ed_verification = st.text_area(UI_LABELS["검증결과"], value=str(rec.get("Verification Result","") or ""), height=100, key="ed_verification")
                ed_remark = st.text_area(UI_LABELS["비고"], value=str(rec.get("Remark","") or ""), height=90, key="ed_remark")
                st.markdown("#### 수정 이력")
                if rec.get("Last Update Summary"):
                    st.info(f"직전 수정 내용: {rec.get('Last Update Summary')}")
                update_summary = st.text_input("이번 수정 내용 / Update Summary / 本次修改内容", placeholder="예: 불량 Sample 확보 및 FACA 원인분석 결과 업데이트")
                if st.button("💾 수정 내용 저장", type="primary", width='stretch'):
                    if not update_summary.strip():
                        st.warning("수정 이력을 위해 '이번 수정 내용'을 입력해주세요.")
                    else:
                        upd = {
                            "Customer": ed_customer, "DRI": ed_dri, "Product": ed_product, "Lot No.": ed_lot,
                            "Process": ed_process, "Failure Category": ed_failure, "Priority": ed_priority,
                            "Status": ed_status, "Response Due": ed_response_due, "FACA Due": ed_faca_due,
                            "FACA No.": ed_faca_no, "Issue Summary": ed_summary, "Customer Request": ed_request,
                            "Internal Action Items": ed_actions, "Required Check Points": ed_checks,
                            "Confirmed Root Cause": ed_root, "Escape Cause": ed_escape,
                            "Corrective Action": ed_corrective, "Verification Result": ed_verification,
                            "Remark": ed_remark, "Close Date": date.today() if ed_status == "Closed" else rec.get("Close Date","")
                        }
                        try:
                            rownum = update_voc(selected_voc, upd, st.session_state.get("current_user","Unknown"), update_summary.strip())
                            st.success(f"✅ {selected_voc} 수정 완료 · Excel {rownum}행 · 백업 생성 완료")
                        except Exception as e:
                            st.error(f"수정 실패: {e}")

with tabs[3]:
    st.subheader("최근 VOC")
    df = read_db()
    q = st.text_input("검색", placeholder="Product / LOT / Customer / Failure")
    show = df.copy()
    if q and len(show):
        mask = show.astype(str).apply(lambda row: row.str.contains(q,case=False,na=False).any(),axis=1)
        show = show[mask]
    st.dataframe(show.tail(100), width='stretch', hide_index=True)

with tabs[4]:
    st.subheader("한 · 중 · 영 변환")
    text = st.text_area("업무 문장", height=180)
    if st.button("🌐 변환"):
        if provider == "Local / No API" or not api_key:
            st.warning("자연어 번역은 AI Provider와 API Key를 연결해야 합니다.")
        else:
            try:
                r = translate_with_provider(text, provider, api_key, model)
                c1,c2,c3 = st.columns(3)
                with c1: st.text_area("한국어",r["ko"],height=200)
                with c2: st.text_area("中文（简体）",r["zh"],height=200)
                with c3: st.text_area("English",r["en"],height=200)
            except Exception as e:
                st.error(str(e))

with tabs[5]:
    st.markdown("""
### V3.0의 핵심
- 기존 VOC 검색 및 수정
- 신규 등록/수정 담당자와 시간 자동 기록
- 수정 전 DB 자동 백업
- Last Update Summary 기록
- 실제 긴 고객 VOC를 **문맥 단위로 분석**
- Product / LOT / QTY / Failure / SPEC / Actual 자동 구조화
- **SPEC vs Actual → 판정**
- **Customer Impact / Priority**
- 고객 요구사항을 독립 Action으로 분해
- Response Due / FACA Due 추출
- 내부 Immediate Action / Check Points / Missing Info 생성
- AI 원인은 항상 **가설**, 확정 원인과 분리
- 한/중/영 요약
- 사람 검토 후 Excel DB 저장

### AI Provider
- **자동 Retry / Fallback**: 503·429·일시적 서버 혼잡 시 같은 모델 재시도 → 같은 Provider의 대체 모델 → Local 규칙 기반 분석 순으로 자동 전환
- **Google Gemini**: 무료 Tier로 테스트하기 좋은 기본 선택
- **OpenRouter Free**: `openrouter/free` 무료 라우터
- **OpenAI**: Luna / Terra / Sol 중 선택
- **Local / No API**: 외부 전송 없이 규칙 기반 fallback

### V2.4 표시 방식
- 화면 항목명: **한국어 / English / 中文** 3개 언어 병기
- AI가 정리한 실무 값: **한국어 우선 표시**
- 제품코드 / LOT / 수량 / 날짜 / SPEC 숫자는 원문 그대로 유지
- 원본 VOC와 한·중·영 요약은 별도 보존

### API 미연결
규칙 기반 fallback이 돌아가지만, 실무형 심층분석 품질은 제한적입니다.

### 보안
개인 API Key를 사용하더라도 회사 데이터를 외부 API로 전송하는 행위 자체는 회사 보안정책의 적용을 받습니다.
""")
